# validator/price_tool.py

import os
import json
import pandas as pd
import re
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import logging

# Importeer de rapporteerfunctie (ervan uitgaande dat die in rapport_utils.py staat)
try:
    from .rapport_utils import genereer_rapport
except ImportError:
    # Fallback voor als het script direct wordt getest (minder relevant voor Streamlit run)
    try:
        from rapport_utils import genereer_rapport
    except ImportError:
        logging.error("Kon genereer_rapport niet importeren. Zorg dat rapport_utils.py bestaat.")
        # Definieer een dummy functie om NameErrors te voorkomen als rapport_utils mist
        def genereer_rapport(*args, **kwargs):
            logging.error("Dummy genereer_rapport aangeroepen omdat import mislukte.")
            print("FOUT: rapport_utils.py of genereer_rapport functie niet gevonden!")
            return None

# -----------------------------
# TEMPLATE-AWARE HELPER FUNCTIES
# -----------------------------

def has_template_generator_stamp(excel_path: str) -> bool:
    """
    Controleert of Excel bestand een Template Generator stamp heeft.
    
    Zoekt naar:
    - _GHX_META sheet 
    - GHX_STAMP named range
    
    Returns:
        True als template generator stamp gevonden
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        
        # Check voor _GHX_META sheet
        has_meta_sheet = "_GHX_META" in wb.sheetnames
        
        # Check voor GHX_STAMP named range
        has_stamp_range = "GHX_STAMP" in wb.defined_names
        
        wb.close()
        
        # Template Generator stamp vereist beide
        return has_meta_sheet and has_stamp_range
        
    except Exception as e:
        logging.warning(f"Fout bij stamp detection: {e}")
        return False

def extract_template_generator_context(excel_path: str) -> Optional[Dict[str, Any]]:
    """
    Extraheert Template Generator context uit _GHX_META sheet.
    
    Returns:
        Dictionary met template context of None bij fout
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        
        if "_GHX_META" not in wb.sheetnames:
            return None
            
        meta_ws = wb["_GHX_META"]
        json_str = meta_ws["A1"].value
        
        if not json_str:
            logging.warning("Geen metadata gevonden in _GHX_META cel A1")
            return None
            
        # Parse JSON metadata
        context = json.loads(json_str)
        
        # Valideer minimum vereiste velden
        required_fields = ['template_choice', 'product_type']
        if not all(field in context for field in required_fields):
            logging.warning(f"Template context mist vereiste velden: {required_fields}")
            return None
            
        wb.close()
        return context
        
    except json.JSONDecodeError as e:
        logging.error(f"Fout bij parsen template context JSON: {e}")
        return None
    except Exception as e:
        logging.error(f"Fout bij extracten template context: {e}")
        return None

def get_context_mandatory_fields(template_config: Dict, context: Dict[str, Any]) -> List[str]:
    """
    Bepaalt mandatory fields gebaseerd op template generator context.
    
    Args:
        template_config: Template configuratie dictionary
        context: Template generator context
        
    Returns:
        Lijst van mandatory field names
    """
    try:
        # Start met GHX standaard mandatory fields
        base_mandatory = template_config.get("default_template", {}).get("mandatory_fields", [])
        mandatory_fields = base_mandatory.copy()
        
        # Haal institution-specific extra mandatory fields op
        institutions = context.get("institutions", [])
        institution_extra = get_institution_mandatory_fields(institutions)
        
        # Voeg extra mandatory fields toe (zonder duplicaten)
        for field in institution_extra:
            if field not in mandatory_fields:
                mandatory_fields.append(field)
        
        logging.info(f"Template Generator mandatory fields: {len(base_mandatory)} GHX + {len(institution_extra)} institutie = {len(mandatory_fields)} totaal")
        
        return mandatory_fields
        
    except Exception as e:
        logging.error(f"Fout bij bepalen context mandatory fields: {e}")
        # Fallback naar default
        return template_config.get("default_template", {}).get("mandatory_fields", [])

def get_institution_mandatory_fields(institutions: List[str]) -> List[str]:
    """
    Geeft extra mandatory fields terug per zorginstelling.
    
    Args:
        institutions: Lijst van instellingscodes (bijv. ['UU', 'Leiden'])
        
    Returns:
        Lijst van extra mandatory field names
    """
    extra_mandatory = []
    
    try:
        # Laad institution rules uit field_validation_v20.json
        validation_config_path = "field_validation_v20.json"
        if os.path.exists(validation_config_path):
            with open(validation_config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            institution_rules = config.get("template_generator", {}).get("institution_mandatory_fields", {})
        else:
            # Fallback als config bestand niet bestaat
            institution_rules = {
                'leids_universitair_medisch_centrum_(lumc,_leiden)': ['Levertijd', 'Duurzaamheid Score', 'CO2 Footprint'],
                'universitair_medisch_centrum_utrecht_(umc_utrecht)': ['Levertijd', 'Milieu Impact Score'],
                'amsterdam_umc_(locaties_amc_en_vumc)': ['Levertijd', 'Circulaire Economie Score'],
                'erasmus_mc': ['Sociale Impact Score'],
                'maastricht_umc+': ['Levertijd', 'Regionale Inkoop Score'],
                'universitair_medisch_centrum_groningen_(umcg)': ['Levertijd', 'Noordelijke Samenwerking Score'],
                # Fallback voor oude codes
                'Leiden': ['Levertijd', 'Duurzaamheid Score', 'CO2 Footprint'],
                'UU': ['Levertijd', 'Milieu Impact Score'],
                'AMC': ['Levertijd', 'Circulaire Economie Score'],
                'Erasmus': ['Sociale Impact Score'],
                'Maastricht': ['Levertijd', 'Regionale Inkoop Score'],
                'Groningen': ['Levertijd', 'Noordelijke Samenwerking Score']
            }
        
        for institution in institutions:
            if institution in institution_rules:
                extra_fields = institution_rules[institution]
                for field in extra_fields:
                    if field not in extra_mandatory:
                        extra_mandatory.append(field)
                        
    except Exception as e:
        logging.error(f"Fout bij laden institution mandatory fields: {e}")
                    
    return extra_mandatory

def get_collapsed_fields(template_context: Dict[str, Any]) -> List[str]:
    """
    Bepaalt welke velden ingeklapt zijn in template generator template.
    
    Deze velden moeten worden weggelaten uit statistieken maar zijn wel aanwezig in Excel.
    
    Returns:
        Lijst van field names die ingeklapt zijn
    """
    collapsed_fields = []
    
    try:
        # Laad collapsed fields configuratie uit field_validation_v20.json
        validation_config_path = "field_validation_v20.json"
        collapsed_config = {}
        
        if os.path.exists(validation_config_path):
            with open(validation_config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            collapsed_config = config.get("template_generator", {}).get("collapsed_fields_by_context", {})
        
        # GS1 velden zijn ingeklapt als gs1_mode = "none" 
        gs1_mode = template_context.get("gs1_mode", "none")
        if gs1_mode == "none":
            gs1_fields = collapsed_config.get("gs1_mode_none", [
                "GTIN", "GLN Fabrikant", "GLN Leverancier", "Doelmarkt Landcode",
                "GS1 Synchronisatie Status", "GDSN Publication Status"
            ])
            collapsed_fields.extend(gs1_fields)
        
        # Staffel-specifieke velden alleen zichtbaar bij staffel templates
        is_staffel = template_context.get("is_staffel_file", False) 
        if not is_staffel:
            staffel_fields = collapsed_config.get("non_staffel", ["Hoeveelheid Van", "Hoeveelheid Tot"])
            collapsed_fields.extend(staffel_fields)
            
        # Product type specifieke velden
        product_type = template_context.get("product_type", "")
        if product_type != "medisch":
            medical_fields = collapsed_config.get("non_medisch", ["GMDN Code", "EMDN Code", "MDR Klasse"])
            collapsed_fields.extend(medical_fields)
            
        if product_type != "lab":
            lab_fields = collapsed_config.get("non_lab", ["Chemische Samenstelling", "Gevaarlijke Stoffen Klasse"])
            collapsed_fields.extend(lab_fields)
            
        # Verwijder duplicaten
        collapsed_fields = list(set(collapsed_fields))
            
        logging.info(f"Ingeklapte velden: {len(collapsed_fields)} velden weggelaten uit statistieken")
        return collapsed_fields
        
    except Exception as e:
        logging.error(f"Fout bij bepalen ingeklapte velden: {e}")
        return []

def determine_mandatory_fields_for_template(excel_path: str) -> List[str]:
    """
    Bepaalt welke velden verplicht zijn voor de gegeven template.
    
    Ondersteunt 3 scenario's:
    1. Template Generator templates (met stamp) - Context-aware mandatory fields
    2. Default templates (geen stamp) - 17 GHX mandatory fields  
    3. Oude leverancier templates - Gebruik aanwezige velden
    """
    try:
        # Laad template configuratie uit field_validation_v20.json
        validation_config_path = "field_validation_v20.json"
        if not os.path.exists(validation_config_path):
            logging.warning(f"Validation config {validation_config_path} niet gevonden, gebruik fallback.")
            return get_fallback_mandatory_fields()
        
        with open(validation_config_path, 'r', encoding='utf-8') as f:
            template_config = json.load(f)
        
        # SCENARIO 3: Template Generator templates
        if has_template_generator_stamp(excel_path):
            logging.info("Template Generator stamp gedetecteerd")
            context = extract_template_generator_context(excel_path)
            
            if context:
                # Template Generator context gevonden
                template_choice = context.get("template_choice", "besteleenheid")
                product_type = context.get("product_type", "facilitair") 
                institutions = context.get("institutions", [])
                
                logging.info(f"Template type: Template Generator - {template_choice} ({product_type})")
                if institutions:
                    logging.info(f"Instellingen: {', '.join(institutions)}")
                
                return get_context_mandatory_fields(template_config, context)
            else:
                logging.warning("Template Generator stamp gevonden maar context extractie gefaald, gebruik default")
        
        # SCENARIO 1: Default template (geen stamp)
        default_config = template_config.get("default_template", {})
        mandatory_fields = default_config.get("mandatory_fields", [])
        
        logging.info(f"Template type: Default (17 verplichte velden)")
        return mandatory_fields
        
    except Exception as e:
        logging.error(f"Fout bij bepalen mandatory fields: {e}")
        return get_fallback_mandatory_fields()

def get_fallback_mandatory_fields() -> List[str]:
    """Fallback lijst van 17 mandatory fields als template config faalt."""
    return [
        "Artikelnummer", "Artikelnaam", "Brutoprijs", "Nettoprijs",
        "Is BestelbareEenheid", "Is BasisEenheid", 
        "Omschrijving Verpakkingseenheid", "UOM Code Verpakkingseenheid",
        "Inhoud Verpakkingseenheid", "UOM Code Basiseenheid", 
        "Inhoud Basiseenheid", "UOM Code Inhoud Basiseenheid",
        "Omrekenfactor", "GHX BTW Code", "UNSPSC",
        "Startdatum Prijs Artikel", "Einddatum Prijs Artikel"
    ]

# -----------------------------
# VALIDATION HELPER FUNCTIES
# -----------------------------

def detect_json_version(validation_config: Dict) -> str:
    """
    Detecteert of we een v18/v19 of v20 JSON structuur hebben.
    
    Returns:
        'v18' voor oude structuur met "fields" object
        'v20' voor nieuwe structuur met "field_validations" object
    """
    if "field_validations" in validation_config:
        return "v20"
    elif "fields" in validation_config:
        return "v18" 
    else:
        raise ValueError("Onbekende JSON structuur: geen 'fields' of 'field_validations' gevonden")

def normalize_v20_to_v18_structure(validation_config: Dict, reference_lists: Dict = None) -> Dict:
    """
    Converteert JSON v2.0 structuur naar v18-compatibele structuur voor backwards compatibility.
    
    Args:
        validation_config: v20 JSON configuratie
        reference_lists: reference_lists.json data (optioneel)
    
    Returns:
        v18-compatibele dictionary
    """
    if reference_lists is None:
        reference_lists = {}
    
    # Start met basis v18 structuur
    v18_structure = {
        "fields": {},
        "red_flags": [],
        "invalid_values": validation_config.get("global_settings", {}).get("null_values", [])
    }
    
    # Voeg error code descriptions toe indien beschikbaar
    if "global_settings" in validation_config and "error_code_descriptions" in validation_config["global_settings"]:
        v18_structure["error_code_descriptions"] = validation_config["global_settings"]["error_code_descriptions"]
    
    # Converteer field_validations naar fields
    field_validations = validation_config.get("field_validations", {})
    
    for field_name, field_config in field_validations.items():
        v18_field = {
            "GHXmandatory": False,
            "read_as_string": field_config.get("data_format") in ["string", "alphanumeric"],
            "allowed_values": []
        }
        
        # Verwerk rules array
        rules = field_config.get("rules", [])
        
        for rule in rules:
            rule_type = rule.get("type")
            condition = rule.get("condition")
            rule_code = rule.get("code")
            
            # Zet mandatory rules om naar GHXmandatory (code 700)
            if condition == "is_empty" and rule_type == "rejection":
                v18_field["GHXmandatory"] = True
            
            # Verwerk list_ref naar allowed_values (code 707)
            if "list_ref" in rule:
                list_ref = rule["list_ref"] 
                if list_ref in reference_lists.get("reference_lists", {}):
                    v18_field["allowed_values"] = reference_lists["reference_lists"][list_ref]
            
            # Verwerk directe allowed_values (code 707)
            elif "allowed_values" in rule:
                v18_field["allowed_values"] = rule["allowed_values"]
            
            # Verwerk value_not_in_list condition naar allowed_values (code 707)
            elif condition == "value_not_in_list" and "params" in rule:
                v18_field["allowed_values"] = rule["params"]
            
            # Bewaar alle rule types voor advanced error handling
            if "validation_rules" not in v18_field:
                v18_field["validation_rules"] = []
            
            v18_field["validation_rules"].append({
                "type": rule_type,
                "condition": condition,
                "code": rule_code,
                "message": rule.get("message", ""),
                "params": rule.get("params", {})
            })
        
        v18_structure["fields"][field_name] = v18_field
    
    # Converteer global_validations naar red_flags  
    global_validations = validation_config.get("global_validations", [])
    
    for validation in global_validations:
        red_flag = {
            "description": validation.get("description", ""),
            "fields": validation.get("fields", []),
            "condition": _map_v20_condition_to_v18(validation.get("condition")),
            "error_message": validation.get("message", ""),
            "code": validation.get("code"),  # Bewaar originele code (800-804)
            "type": validation.get("type", "flag"),  # Bewaar type info
            "id": validation.get("id")  # Bewaar unique ID
        }
        v18_structure["red_flags"].append(red_flag)
    
    return v18_structure

def _map_v20_condition_to_v18(v20_condition: str) -> str:
    """Maps v20 condition names naar v18 condition names."""
    mapping = {
        "all_fields_empty": "both_empty",
        "uom_relation_conflict": "uom_relation", 
        "uom_match_if_base_and_orderable": "uom_match",
        "content_match_if_base_and_orderable": "content_match",
        "template_column_missing": "template_check",
        "uom_description_format_mismatch": "uom_description_format",
        "incomplete_set": "incomplete_dimensions"
    }
    return mapping.get(v20_condition, v20_condition)

def normalize_template_header(header: str) -> str:
    """
    Normaliseert nieuwe GHX template headers naar hun standaard Nederlandse vorm.
    
    Template headers hebben dit formaat:
    Nederlandse Naam
    (extra info)
    
    _________________
    English Name
    (extra info)
    
    We willen alleen: "Nederlandse Naam"
    """
    if not isinstance(header, str):
        return ''
    
    # Vervang non-breaking spaces
    normalized = header.replace('\xa0', ' ')
    
    # Split op newlines
    lines = [line.strip() for line in normalized.split('\n') if line.strip()]
    
    if not lines:
        return ''
    
    # Eerste regel is de Nederlandse naam
    dutch_name = lines[0].strip()
    
    # Verwijder eventuele haakjes van de eerste regel (we willen alleen de basis naam)
    if '(' in dutch_name:
        dutch_name = dutch_name.split('(')[0].strip()
    
    return dutch_name

def clean_column_name(col: str) -> str:
    """Schoont een kolomnaam op voor mapping vergelijking."""
    if not isinstance(col, str):
        return ''
    
    # Vervang non-breaking spaces met normale spaces
    cleaned_col = col.replace('\xa0', ' ')
    
    # Split op newlines en verwerk alle regels
    lines = cleaned_col.split('\n')
    first_line = lines[0].strip()
    
    # Zoek naar haakjes op volgende regels en voeg toe aan eerste regel
    for line in lines[1:]:
        line = line.strip()
        if line.startswith('(') and line.endswith(')'):
            # Voeg haakjes toe aan eerste regel als het nog niet aanwezig is
            if line not in first_line:
                first_line = first_line.rstrip() + ' ' + line
            break  # Neem alleen de eerste haakjes
    
    # Verwijder eventuele underscore-delen, maar alleen als het niet begint met underscore
    if '_' in first_line and not first_line.startswith('_'):
        first_line = first_line.split('_')[0].strip()
    
    return first_line.lower()

def clean_supplier_header(header: str) -> str:
    """Clean supplier header by extracting Dutch name before dash."""
    if not header:
        return ""
    # Verwijder newlines en splits op dash om Nederlandse naam vóór de dash te behouden
    clean = str(header).split('\n')[0].strip()
    # Split op " - " en neem het Nederlandse deel (vóór de dash)
    if " - " in clean:
        clean = clean.split(" - ")[0].strip()
    return clean

def map_headers(df: pd.DataFrame, mapping_config: Dict, return_mapping: bool = False) -> Tuple[pd.DataFrame, List[str], Dict[str, str], Dict[str, str]] | Tuple[pd.DataFrame, List[str], Dict[str, str]]:
    """Mapt de headers van het DataFrame naar de GHX standaard headers."""
    # Haal mapping uit configuratie - check of we de volledige config of alleen standard_headers krijgen
    if "standard_headers" in mapping_config:
        header_mapping = {k: v["alternatives"] for k, v in mapping_config["standard_headers"].items()}
    else:
        # We krijgen direct de standard_headers dictionary
        header_mapping = {k: v["alternatives"] for k, v in mapping_config.items()}

    # DEBUG: Log de header schoonmaak stap voor stap (debug level)
    logging.debug("=== STAP 2: HEADER SCHOONMAAK ===")
    
    # Maak een dictionary van originele kolom -> opgeschoonde kolomnaam
    cleaned_columns = {}
    for col in df.columns:
        cleaned = clean_column_name(col)
        cleaned_columns[col] = cleaned
        # Log problematische headers
        if "hoogte" in str(col).lower() or "inhoud basiseenheid" in str(col).lower():
            logging.debug(f"SCHOONMAAK - Origineel: {repr(col)} → Schoon: {repr(cleaned)}")

    # Maak een reverse mapping: opgeschoonde alternatieve naam -> standaard header
    reverse_mapping = {}
    for std_header, alternatives in header_mapping.items():
        # Voeg de standaard header zelf ook toe als alternatief (lowercase, schoon)
        cleaned_std_header = clean_column_name(std_header)
        if cleaned_std_header not in reverse_mapping: # Voorkom overschrijven door alternatief
             reverse_mapping[cleaned_std_header] = std_header
        # Voeg alternatieven toe
        for alt in alternatives:
            cleaned_alt = clean_column_name(alt)
            reverse_mapping[cleaned_alt] = std_header # Alternatief wijst naar standaard
    
    # AUTOMATISCHE FALLBACK: Voor UOM headers zonder haakjes, zoek naar equivalente met haakjes
    uom_fallback_mapping = {}
    for std_header in header_mapping.keys():
        if "(UOM)" in std_header:
            # Maak fallback zonder (UOM)
            base_name = std_header.replace(" (UOM)", "").strip()
            base_name_clean = clean_column_name(base_name)
            if base_name_clean not in reverse_mapping:
                uom_fallback_mapping[base_name_clean] = std_header
                logging.debug(f"UOM fallback: {repr(base_name_clean)} → {std_header}")
    
    # Voeg fallback mappings toe
    reverse_mapping.update(uom_fallback_mapping)

    mapped_columns = {}
    unrecognized = []
    duplicates = {} # Houd bij welke standaard header duplicaten heeft en wat de originelen waren
    original_column_mapping = {} # Houd bij welke originele kolom naar welke (niet-duplicate) standaard header mapt

    # Houd bij welke standaard headers we al hebben toegewezen om duplicaten te nummeren
    assigned_std_headers_count = {}

    # DEBUG: Log de reverse mapping voor problematische headers (debug level)
    logging.debug("=== STAP 3: MAPPING VERGELIJKING ===")
    
    for original_col in df.columns:
        clean_col = cleaned_columns[original_col]
        std_header = reverse_mapping.get(clean_col) # Zoek standaard header
        
        # Debug logging voor problematische headers
        if "hoogte" in str(original_col).lower() or "inhoud basiseenheid" in str(original_col).lower():
            logging.debug(f"MAPPING - Origineel: {repr(original_col)}")
            logging.debug(f"MAPPING - Schoon: {repr(clean_col)}")
            logging.debug(f"MAPPING - Gevonden std_header: {std_header}")
            # Toon beschikbare alternatieven die zouden kunnen matchen
            for std_h, alternatives in header_mapping.items():
                for alt in alternatives:
                    alt_clean = clean_column_name(alt)
                    if alt_clean == clean_col:
                        logging.debug(f"MAPPING - MATCH GEVONDEN: {std_h} via alternatief {repr(alt)} → {repr(alt_clean)}")

        if std_header:
            # Standaard header gevonden
            count = assigned_std_headers_count.get(std_header, 0)
            if count == 0:
                # Eerste keer dat we deze standaard header zien
                new_header = std_header
                mapped_columns[original_col] = new_header
                original_column_mapping[new_header] = original_col # Map standaard naar origineel
                assigned_std_headers_count[std_header] = 1
            else:
                # Dit is een duplicaat
                count += 1
                new_header = f"{std_header}_DUPLICAAT_{count}"
                duplicates.setdefault(std_header, []).append(original_col) # Noteer origineel bij duplicaat
                mapped_columns[original_col] = new_header # Hernoem kolom in df
                assigned_std_headers_count[std_header] = count
        else:
            # Geen standaard header gevonden, gebruik originele (opgeschoonde?) naam
            # We gebruiken de *originele* naam in mapped_columns om te zorgen dat rename werkt
            mapped_columns[original_col] = original_col
            if not str(original_col).lower().startswith('algemeen'): # Negeer 'algemeen' kolom
                 unrecognized.append(original_col) # Noteer als onherkend (indien niet 'algemeen')

    # Hernoem de kolommen in het DataFrame
    df = df.rename(columns=mapped_columns)

    # Maak dictionary met details over duplicaten (welke _DUPLICAAT_X hoort bij welk origineel)
    duplicate_headers_details = {}
    for header, original_cols_list in duplicates.items():
        # Start nummering duplicaten vanaf _2
        for i, original_col_name in enumerate(original_cols_list):
             duplicate_name = f"{header}_DUPLICAAT_{i+2}"
             duplicate_headers_details[duplicate_name] = original_col_name

    if return_mapping:
        # Geef df, onherkende lijst, details duplicaten, en origineel->standaard mapping terug
        return df, unrecognized, duplicate_headers_details, original_column_mapping
    else:
        # Geef alleen df, onherkende lijst, en details duplicaten terug
        return df, unrecognized, duplicate_headers_details

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Schoont het DataFrame op (logica uit Code 3)."""
    # Check eerst of 'Artikelnummer' kolom bestaat (na mapping)
    if 'Artikelnummer' not in df.columns:
        logging.error("'Artikelnummer' kolom niet gevonden na mapping. Opschonen gestopt.")
        # Overweeg hier een error te raisen of de originele df terug te geven
        # raise ValueError("'Artikelnummer' kolom niet gevonden na mapping.")
        return df # Geeft df terug zoals het was

    # Verwijder rijen waar alles leeg is
    df = df.dropna(how='all')

    # Vind eerste en laatste rij met een geldig Artikelnummer
    first_valid = df['Artikelnummer'].first_valid_index()
    last_valid = df['Artikelnummer'].last_valid_index()

    # Beperk DataFrame tot deze rijen
    if first_valid is not None and last_valid is not None:
        logging.info(f"Data beperkt tot rijen (0-based index): {first_valid} t/m {last_valid}")
        df = df.loc[first_valid:last_valid].reset_index(drop=True)
    else:
        logging.warning("Geen geldige 'Artikelnummer' waarden gevonden, dataframe is mogelijk leeg.")
        return pd.DataFrame(columns=df.columns) # Geeft lege dataframe terug met dezelfde kolommen

    # Verwijder uitleg/instructie rij (check op eerste rij na beperking)
    if not df.empty:
        first_value = str(df.loc[0, 'Artikelnummer']).strip()
        explanation_keywords = [
            "karakters", "characters", "the unique", "explanation", "unique article number",
            "instructie", "instruction", "you, as the", "vul het unieke",
            "assign to", "supplier", "artikelnummer dat u als leverancier"
        ]
        # Check op lengte of keywords
        if (len(first_value) > 70 or # Iets kortere check
            any(keyword in first_value.lower() for keyword in explanation_keywords)):
            logging.info(f"Mogelijke uitleg/instructie rij verwijderd op basis van inhoud: '{first_value[:100]}...'")
            df = df.iloc[1:].reset_index(drop=True)

        # Extra check voor voorbeeld-/demo-rij (op de *nieuwe* eerste rij)
        if not df.empty:
            first_value_after_potential_removal = str(df.loc[0, 'Artikelnummer']).strip()
            voorbeeldwaarden = ["ghx-12345", "voorbeeld", "demo", "example", "test"] # Uitgebreide lijst
            if first_value_after_potential_removal.lower() in voorbeeldwaarden:
                logging.info(f"Mogelijke voorbeeld/demo rij verwijderd op basis van inhoud: '{first_value_after_potential_removal}'")
                df = df.iloc[1:].reset_index(drop=True)

    # Verwijder 'algemeen' kolom indien die nog bestaat (zou niet moeten na mapping, maar voor zekerheid)
    if 'algemeen' in df.columns:
         logging.warning("Kolom 'algemeen' gevonden en verwijderd na mapping.")
         df = df.drop(columns=['algemeen'])
    # Check ook op variaties met hoofdletters etc.
    algemeen_cols = [col for col in df.columns if str(col).strip().lower() == 'algemeen']
    if algemeen_cols:
        logging.warning(f"Kolommen {algemeen_cols} gevonden die lijken op 'algemeen' en worden verwijderd.")
        df = df.drop(columns=algemeen_cols)


    return df

def validate_field(field_name: str, value: Any, field_rules: dict, invalid_values: list, row_data: dict = None) -> list:
    """Valideert een enkele waarde op basis van de regels (uit Code 4)."""
    errors = []
    value_str = str(value).strip() # Werk met de string representatie

    # 1. Leeg check (incl. invalid values)
    is_empty_or_invalid = pd.isnull(value) or value_str == '' or value_str.lower() in invalid_values
    if is_empty_or_invalid:
        # Is het veld verplicht (GHXmandatory of via dependency)?
        is_required = field_rules.get("GHXmandatory", False)
        # Check dependency rule
        if not is_required and "depends_on" in field_rules and row_data:
             dependency = field_rules["depends_on"]
             dep_field = dependency.get("fields", [""])[0]
             dep_cond = dependency.get("condition")
             dep_value = str(row_data.get(dep_field, "")).strip()
             # Als conditie '1' is en de afhankelijke waarde is '1', dan is dit veld ook verplicht
             if dep_cond == "1" and dep_value == "1":
                 is_required = True

        if is_required:
            error_message = field_rules.get("error_messages", {}).get("required", f"Veld '{field_name}' is verplicht maar leeg.")
            error_code = field_rules.get("error_codes", {}).get("required", "700") # Default code voor leeg verplicht
            errors.append({'message': error_message, 'code': error_code})
        return errors # Stop validatie als leeg/invalid

    # 2. String checks (alleen als niet leeg/invalid)
    if field_rules.get("type") == "string":
        min_length = field_rules.get("min_length")
        if min_length and len(value_str) < min_length:
            error_message = field_rules.get("error_messages", {}).get("min_length", f"Waarde '{value_str[:20]}...' is te kort (min {min_length}).")
            error_code = field_rules.get("error_codes", {}).get("min_length", "701")
            errors.append({'message': error_message, 'code': error_code})
        max_length = field_rules.get("max_length")
        if max_length and len(value_str) > max_length:
            error_message = field_rules.get("error_messages", {}).get("max_length", f"Waarde '{value_str[:20]}...' is te lang (max {max_length}).")
            error_code = field_rules.get("error_codes", {}).get("max_length", "702")
            errors.append({'message': error_message, 'code': error_code})
        allowed = field_rules.get("allowed_values")
        if allowed:
            # Normaliseer zowel de waarde als de toegestane waarden voor vergelijking
            normalized_value = value_str.upper()
            allowed_normalized = [str(x).strip().upper() for x in allowed]
            if normalized_value not in allowed_normalized:
                error_message = field_rules.get("error_messages", {}).get("allowed_values", f"Waarde '{value_str}' niet toegestaan. Moet zijn: {', '.join(map(str,allowed))}.")
                error_code = field_rules.get("error_codes", {}).get("allowed_values", "707")
                errors.append({'message': error_message, 'code': error_code})

    # 3. Numeric checks (alleen als niet leeg/invalid)
    if field_rules.get("type") == "numeric":
        # Verwijder duizendtalscheidingstekens voor conversie
        cleaned_value_str = value_str.replace('.', '') # Veronderstelt '.' als duizendtal, pas aan indien nodig
        cleaned_value_str = cleaned_value_str.replace(field_rules.get("decimal_separator", ","), ".") # Vervang decimaal door '.'

        try:
            numeric_value = float(cleaned_value_str)
            # Check max integer digits
            max_int = field_rules.get("max_integer_digits")
            if max_int is not None:
                # Gebruik abs() voor negatieve getallen
                integer_part_str = str(int(abs(numeric_value)))
                if len(integer_part_str) > max_int:
                    error_message = field_rules.get("error_messages", {}).get("max_integer_digits", f"Te veel cijfers voor de komma ({len(integer_part_str)} > {max_int}).")
                    error_code = field_rules.get("error_codes", {}).get("max_integer_digits", "705")
                    errors.append({'message': error_message, 'code': error_code})
            # Check max decimal digits (op basis van originele string na vervanging separator)
            max_dec = field_rules.get("max_decimal_digits")
            if max_dec is not None:
                original_parts = value_str.split(field_rules.get("decimal_separator", ","))
                if len(original_parts) == 2:
                     # Verwijder eventuele spaties of duizendtal-tekens uit decimaal deel
                     decimal_part_cleaned = re.sub(r'[^\d]', '', original_parts[1])
                     # Tel alleen relevante decimalen
                     relevant_decimals = len(decimal_part_cleaned.rstrip('0'))
                     if relevant_decimals > max_dec:
                          error_message = field_rules.get("error_messages", {}).get("max_decimal_digits", f"Te veel decimalen ({len(decimal_part_cleaned)} > {max_dec}).")
                          error_code = field_rules.get("error_codes", {}).get("max_decimal_digits","706")
                          errors.append({'message': error_message, 'code': error_code})
        except (ValueError, TypeError):
            error_message = field_rules.get("error_messages", {}).get("numeric", f"Waarde '{value_str}' is geen geldig getal.")
            error_code = field_rules.get("error_codes", {}).get("numeric", "704")
            errors.append({'message': error_message, 'code': error_code})

    # 4. Regex / invalid_format validatie (alleen als niet leeg/invalid)
    # Let op: Regex wordt toegepast op de originele string waarde
    if "validation" in field_rules and "regex" in field_rules["validation"]:
        pattern = field_rules["validation"]["regex"]
        # Voeg ankers toe als ze ontbreken voor volledige match
        if not pattern.startswith('^'): pattern = '^' + pattern
        if not pattern.endswith('$'): pattern = pattern + '$'
        try:
            if not re.match(pattern, value_str):
                error_message = field_rules.get("error_messages", {}).get("invalid_format", f"Waarde '{value_str}' heeft ongeldig formaat.")
                error_code = field_rules.get("error_codes", {}).get("invalid_format", "722")
                errors.append({'message': error_message, 'code': error_code})
        except re.error as e:
             logging.warning(f"Ongeldige regex voor veld {field_name}: {pattern} - Fout: {e}")


    # 5. Overige validaties (zoals mismatch, etc.)
    # Voorbeeld mismatch validatie (uit Code 4) - NOG NIET GEIMPLEMENTEERD IN DEZE FUNCTIE
    # if "depends_on" in field_rules and row_data:
    #     depends_on = field_rules["depends_on"]
    #     if depends_on.get("condition") == "mismatch":
    #         # Implementeer logica om velden te vergelijken
    #         # fields_to_compare = depends_on.get("fields", [])
    #         # if field_name in fields_to_compare and len(fields_to_compare) > 1: ...
    #         error_message = depends_on.get("error_message", "Waarden komen niet overeen")
    #         error_code = field_rules.get("error_codes", {}).get("mismatch", "81") # Voorbeeld code
    #         errors.append({'message': error_message, 'code': error_code})

    return errors

def validate_uom_relationships(df: pd.DataFrame, validation_results: list, validation_config: dict, original_column_mapping: dict) -> list:
    """
    Controleert relaties tussen UOM Codes, Is BestelbareEenheid en Is BasisEenheid.
    (Grotendeels overgenomen uit Code 4)
    """
    required_cols = [
        "Is BestelbareEenheid", "Is BasisEenheid",
        "UOM Code Verpakkingseenheid", "UOM Code Basiseenheid",
        "Inhoud Verpakkingseenheid"
    ]

    # Als cruciale kolommen ontbreken, sla over (oude template?)
    if not all(col in df.columns for col in required_cols):
        logging.warning("Benodigde UOM-kolommen niet allemaal aanwezig, UOM-relatie validatie overgeslagen.")
        return validation_results

    uom_relation_errors_found = False # Flag om te zien of we code 724 fouten toevoegen
    uom_red_flag_config = None
    uom_description_flag_config = None
    
    # Zoek naar beide red flag configs - support v18 and v20
    flags_to_search = validation_config.get("red_flags", []) if "red_flags" in validation_config else validation_config.get("global_validations", [])
    
    for flag_config in flags_to_search:
         condition = flag_config.get("condition")
         if condition in ["uom_relation", "uom_relation_conflict"]:
             uom_red_flag_config = flag_config
         elif condition in ["uom_description_format", "uom_description_format_mismatch"]:
             uom_description_flag_config = flag_config

    omschrijving_format_mismatch_count = 0
    rij_offset = 3 # Start rijnummer in Excel (1-based) na header(s) en instructie(s)

    for idx, row in df.iterrows():
        excel_row_num = idx + rij_offset # Werkelijke rijnummer in Excel voor meldingen
        try:
            # Gebruik .get() met default lege string voor robuustheid
            is_besteleenheid = str(row.get("Is BestelbareEenheid", "")).strip()
            is_basiseenheid = str(row.get("Is BasisEenheid", "")).strip()
            uom_verpakking = str(row.get("UOM Code Verpakkingseenheid", "")).strip()
            uom_basis = str(row.get("UOM Code Basiseenheid", "")).strip()
            inhoud_verpakking_str = str(row.get("Inhoud Verpakkingseenheid", "")).strip()
            inhoud_verpakking = None
            if inhoud_verpakking_str and inhoud_verpakking_str.lower() not in ["nan", "none", ""]:
                try:
                    # Vervang eerst komma door punt voor conversie
                    inhoud_verpakking = float(inhoud_verpakking_str.replace(",", "."))
                except (ValueError, TypeError):
                    # Als conversie faalt, log het eventueel maar ga door.
                    # De standaard validatie pikt dit mogelijk al op als 'geen geldig getal'.
                    logging.debug(f"Kon 'Inhoud Verpakkingseenheid' ({inhoud_verpakking_str}) niet converteren naar float op rij {excel_row_num}")
                    pass # Blijft None als conversie faalt

            # --- UOM Checks ---
            supplier_col_map = { # Map naar typische supplier kolom namen indien beschikbaar
                 "Is BestelbareEenheid": original_column_mapping.get("Is BestelbareEenheid", ""),
                 "Is BasisEenheid": original_column_mapping.get("Is BasisEenheid", ""),
                 "UOM Code Verpakkingseenheid": original_column_mapping.get("UOM Code Verpakkingseenheid", ""),
                 "UOM Code Basiseenheid": original_column_mapping.get("UOM Code Basiseenheid", ""),
                 "Inhoud Verpakkingseenheid": original_column_mapping.get("Inhoud Verpakkingseenheid", "")
            }


            # CHECK 1 & 2: Beide 1
            if is_besteleenheid == "1" and is_basiseenheid == "1":
                # Check 1: UOMs gelijk?
                if uom_verpakking and uom_basis and uom_verpakking != uom_basis:
                    uom_relation_errors_found = True
                    validation_results.append({
                        "Rij": excel_row_num, "GHX Kolom": "UOM Code Verpakkingseenheid",
                        "Supplier Kolom": clean_supplier_header(supplier_col_map["UOM Code Verpakkingseenheid"]),
                        "Veldwaarde": uom_verpakking,
                        "Foutmelding": f"CONFLICT: Als IsBestelbaar=1 én IsBasis=1, moeten UOMs gelijk zijn (nu: '{uom_verpakking}' vs '{uom_basis}').",
                        "code": "724"
                    })
                # Check 2: Inhoud = 1?
                if inhoud_verpakking is not None and inhoud_verpakking != 1:
                    uom_relation_errors_found = True
                    validation_results.append({
                        "Rij": excel_row_num, "GHX Kolom": "Inhoud Verpakkingseenheid",
                        "Supplier Kolom": clean_supplier_header(supplier_col_map["Inhoud Verpakkingseenheid"]),
                        "Veldwaarde": inhoud_verpakking_str,
                        "Foutmelding": "CONFLICT: Als IsBestelbaar=1 én IsBasis=1, moet Inhoud Verpakkingseenheid '1' zijn.",
                        "code": "724"
                    })

            # CHECK 3 & 4: Verschillend (en beide ingevuld)
            elif (is_besteleenheid in ["1", "0"] and is_basiseenheid in ["1", "0"] and 
                  is_besteleenheid != is_basiseenheid):
                # Check 3: UOMs verschillend?
                if uom_verpakking and uom_basis and uom_verpakking == uom_basis:
                     uom_relation_errors_found = True
                     validation_results.append({
                         "Rij": excel_row_num, "GHX Kolom": "UOM Code Verpakkingseenheid",
                         "Supplier Kolom": clean_supplier_header(supplier_col_map["UOM Code Verpakkingseenheid"]),
                         "Veldwaarde": uom_verpakking,
                         "Foutmelding": "CONFLICT: Als IsBestelbaar/IsBasis verschillend zijn, moeten UOMs ook verschillend zijn.",
                         "code": "724"
                     })
                # Check 4: Inhoud != 1?
                if inhoud_verpakking is not None and inhoud_verpakking == 1:
                    uom_relation_errors_found = True
                    validation_results.append({
                        "Rij": excel_row_num, "GHX Kolom": "Inhoud Verpakkingseenheid",
                        "Supplier Kolom": clean_supplier_header(supplier_col_map["Inhoud Verpakkingseenheid"]),
                        "Veldwaarde": inhoud_verpakking_str,
                        "Foutmelding": "CONFLICT: Als IsBestelbaar/IsBasis verschillend zijn, mag Inhoud Verpakkingseenheid geen '1' zijn.",
                        "code": "724"
                    })

            # CHECK 5: Omschrijving Verpakkingseenheid consistentie (als Red Flag / Waarschuwing)
            if "Omschrijving Verpakkingseenheid" in df.columns:
                 omschrijving = str(row.get("Omschrijving Verpakkingseenheid", "")).strip()
                 if omschrijving and uom_verpakking and inhoud_verpakking is not None:
                     # Gebruik originele string voor inhoud check, maar zonder '.0' als het een heel getal was
                     inhoud_str_check = inhoud_verpakking_str
                     if inhoud_str_check.endswith('.0'):
                          inhoud_str_check = inhoud_str_check[:-2]

                     # Basis check: bevat omschrijving (ongeacht hoofdletters) de inhoud en de UOM code?
                     if not (inhoud_str_check.lower() in omschrijving.lower() and uom_verpakking.lower() in omschrijving.lower()):
                        omschrijving_format_mismatch_count += 1
                        # We voegen de Red Flag pas na de loop toe als er mismatches zijn

        except Exception as e:
            logging.error(f"Fout tijdens UOM validatie logica voor Excel rij {excel_row_num}: {e}")
            continue

    # --- Na de loop ---

    # Voeg geconsolideerde Red Flag toe voor Omschrijving Verpakkingseenheid mismatch
    if omschrijving_format_mismatch_count > 0 and uom_description_flag_config:
         # Gebruik bericht uit JSON config - v20 gebruikt 'message', v18 gebruikt 'error_message'
         json_omschrijving_message = (uom_description_flag_config.get("message") or 
                                      uom_description_flag_config.get("error_message") or
                                      "Verschillende 'Omschrijving Verpakkingseenheid' velden komen mogelijk niet overeen met de verwachte notatie. Controleer of deze velden de juiste UOM code bevatten.")
         validation_results.append({
             "Rij": 0, "GHX Kolom": "RED FLAG",
             "Supplier Kolom": "Omschrijving Verpakkingseenheid",
             "Veldwaarde": "",
             "Foutmelding": json_omschrijving_message,
             "code": uom_description_flag_config.get("code", "721")
         })

    # Voeg Red Flag toe als er UOM-relatie fouten (code 724) waren EN er een config voor is
    if uom_relation_errors_found and uom_red_flag_config:
        # v20 gebruikt 'message', v18 gebruikt 'error_message'
        message = uom_red_flag_config.get("message") or uom_red_flag_config.get("error_message")
        if message:
            # Voorkom duplicaten van deze specifieke red flag
            flag_exists = any(r.get("GHX Kolom") == "RED FLAG" and r.get("Foutmelding") == message for r in validation_results)
            if not flag_exists:
                validation_results.append({
                    "Rij": 0, "GHX Kolom": "RED FLAG", "Supplier Kolom": "", "Veldwaarde": "",
                    "Foutmelding": message,
                    "code": "" # Geen specifieke code voor deze algemene vlag
                })

    return validation_results


def validate_field_v20_native(field_name: str, value: Any, field_config: dict, invalid_values: list, row_data: dict = None, reference_lists: dict = None) -> list:
    """Native v20 validatie zonder conversie naar v18."""
    errors = []
    value_str = str(value).strip() if not pd.isnull(value) else ""
    
    # Check if empty/invalid
    is_empty_or_invalid = pd.isnull(value) or value_str == '' or value_str.lower() in invalid_values
    
    # Process validation rules from v20 structure
    rules = field_config.get("rules", [])
    
    for rule in rules:
        rule_type = rule.get("type")  # rejection, correction, flag
        condition = rule.get("condition")
        code = rule.get("code")
        message = rule.get("message", f"Validatie fout in veld '{field_name}'")
        params = rule.get("params")
        
        should_trigger = False
        
        # Handle different validation conditions
        if condition == "is_empty":
            should_trigger = is_empty_or_invalid
            
        elif condition == "is_not_numeric" and not is_empty_or_invalid:
            # Check if value is numeric
            try:
                float(value_str)
                should_trigger = False
            except (ValueError, TypeError):
                should_trigger = True
                
        elif condition == "value_not_in_list" and not is_empty_or_invalid:
            # Check for direct params list
            if isinstance(params, list):
                normalized_value = value_str.upper()
                allowed_normalized = [str(x).strip().upper() for x in params]
                should_trigger = normalized_value not in allowed_normalized
            # Check for list_ref
            elif "list_ref" in rule and reference_lists:
                list_ref = rule["list_ref"]
                if list_ref in reference_lists.get("reference_lists", {}):
                    allowed_values = reference_lists["reference_lists"][list_ref]
                    normalized_value = value_str.upper()
                    allowed_normalized = [str(x).strip().upper() for x in allowed_values]
                    should_trigger = normalized_value not in allowed_normalized
                
        elif condition == "min_length" and not is_empty_or_invalid:
            if isinstance(params, (int, float)):
                should_trigger = len(value_str) < params
                
        elif condition == "max_length" and not is_empty_or_invalid:
            if isinstance(params, (int, float)):
                should_trigger = len(value_str) > params
                
        elif condition == "is_empty_when_dependency_filled" and is_empty_or_invalid and row_data:
            # Check if dependency field is filled
            if isinstance(params, list) and len(params) > 0:
                dependency_field = params[0]
                dependency_value = row_data.get(dependency_field)
                dependency_str = str(dependency_value).strip() if not pd.isnull(dependency_value) else ""
                dependency_filled = dependency_str != '' and dependency_str.lower() not in invalid_values
                should_trigger = dependency_filled
                
        elif condition == "mismatch_calculation" and not is_empty_or_invalid and row_data:
            # Check calculation: veld1 * veld2 = current field
            if isinstance(params, list) and len(params) >= 3:
                field1_name = params[0]  # "Inhoud Verpakkingseenheid"
                operator = params[1]     # "*"
                field2_name = params[2]  # "Inhoud Basiseenheid"
                
                field1_value = row_data.get(field1_name)
                field2_value = row_data.get(field2_name)
                
                try:
                    current_value = float(value_str)
                    val1 = float(field1_value) if field1_value not in [None, '', 'nan'] else None
                    val2 = float(field2_value) if field2_value not in [None, '', 'nan'] else None
                    
                    if val1 is not None and val2 is not None:
                        if operator == "*":
                            calculated_value = val1 * val2
                        elif operator == "/":
                            calculated_value = val1 / val2 if val2 != 0 else None
                        else:
                            calculated_value = None
                            
                        if calculated_value is not None:
                            # Allow kleine afwijkingen door floating point precision
                            should_trigger = abs(current_value - calculated_value) > 0.01
                except (ValueError, TypeError, ZeroDivisionError):
                    should_trigger = False  # Kan berekening niet uitvoeren
        
        elif condition == "is_not_boolean" and not is_empty_or_invalid:
            # Check if value is valid boolean (1, 0, ja, nee, yes, no, true, false)
            valid_booleans = ["1", "0", "ja", "nee", "yes", "no", "true", "false"]
            should_trigger = value_str.lower() not in valid_booleans
            
        elif condition == "is_not_exact_length_numeric" and not is_empty_or_invalid:
            # Check exact length for numeric strings (e.g., UNSPSC codes)
            if isinstance(params, (int, float)):
                should_trigger = len(value_str) != params or not value_str.isdigit()
                
        elif condition == "is_duplicate_artikelnummer" and not is_empty_or_invalid and row_data:
            # This would require additional context about other rows
            # For now, we'll skip this complex validation
            should_trigger = False
            
        elif condition == "invalid_au_risk_combination" and not is_empty_or_invalid and row_data:
            # Cross-field validation: AU code + Risk class combination
            au_field = "Code voor Aanvullende Productclassificatie"
            au_value = str(row_data.get(au_field, "")).strip()
            
            if au_value == "76":
                # MDR/IVDR valid risk classes
                valid_risk_classes = ["EU_CLASS_I", "EU_CLASS_IIA", "EU_CLASS_IIB", "EU_CLASS_III", 
                                    "EU_CLASS_A", "EU_CLASS_B", "EU_CLASS_C", "EU_CLASS_D"]
                should_trigger = value_str.upper() not in [rc.upper() for rc in valid_risk_classes]
            elif au_value == "85":
                # MDD/AIMDD/IVDD valid risk classes
                valid_risk_classes = ["EU_CLASS_I", "EU_CLASS_IIA", "EU_CLASS_IIB", "EU_CLASS_III",
                                    "IVDD_ANNEX_II_LIST_A", "IVDD_ANNEX_II_LIST_B", 
                                    "IVDD_DEVICES_SELF_TESTING", "IVDD_GENERAL"]
                should_trigger = value_str.upper() not in [rc.upper() for rc in valid_risk_classes]
            else:
                should_trigger = False
                
        elif condition == "medical_product_missing_classification" and is_empty_or_invalid and row_data:
            # UNSPSC 42xxxx requires GMDN or EMDN
            unspsc_field = "UNSPSC Code"
            unspsc_value = str(row_data.get(unspsc_field, "")).strip()
            
            if unspsc_value.startswith("42") and len(unspsc_value) >= 2:
                # Check if GMDN or EMDN is filled
                gmdn_value = str(row_data.get("GMDN Code", "")).strip()
                emdn_value = str(row_data.get("EMDN Code", "")).strip()
                
                gmdn_filled = gmdn_value != '' and gmdn_value.lower() not in invalid_values
                emdn_filled = emdn_value != '' and emdn_value.lower() not in invalid_values
                
                # Trigger if current field is empty and NEITHER GMDN nor EMDN is filled
                should_trigger = not gmdn_filled and not emdn_filled
            else:
                should_trigger = False
            
        # Add more conditions as needed...
        
        if should_trigger:
            errors.append({
                'message': message,
                'code': str(code),
                'type': rule_type
            })
    
    return errors

def validate_dataframe(df: pd.DataFrame, validation_config: dict, original_column_mapping: dict, template_context: Dict[str, Any] = None) -> Tuple[list, dict, list, dict]:
    """
    Valideert het DataFrame en berekent template-aware statistieken.
    
    Args:
        df: DataFrame om te valideren
        validation_config: Validatie configuratie
        original_column_mapping: Kolom mapping
        template_context: Template Generator context (None voor default templates)
    
    Retourneert: (results, filled_percentages, red_flag_messages, errors_per_field)
    """
    results = [] # Lijst met alle individuele fout/warning dicts
    
    # Detect v20 vs v18 structure
    if "field_validations" in validation_config:
        # v20 structure
        fields_config = validation_config.get("field_validations", {})
        invalid_values_config = validation_config.get("global_settings", {}).get("null_values", [])
    else:
        # v18 structure
        fields_config = validation_config.get("fields", {})
        invalid_values_config = validation_config.get("invalid_values", [])
    invalid_values = [str(val).lower() for val in invalid_values_config]
    filled_counts = {} # Houdt telling bij per veld
    field_validation_results = {} # Houdt per veld een lijst van error dicts bij
    red_flag_messages_list = [] # Verzamelt Red Flag berichten met codes: [{"message": str, "code": str}, ...]
    total_rows = len(df)
    rij_offset = 3 # Start rijnummer in Excel na header(s)/instructie(s)

    # Bepaal ingeklapte velden voor Template Generator templates
    collapsed_fields = []
    if template_context:
        collapsed_fields = get_collapsed_fields(template_context)
        logging.info(f"Template-aware statistieken: {len(collapsed_fields)} ingeklapte velden uitgesloten")
    
    # Initialiseer summary_stats (zal worden geretourneerd als filled_percentages)
    summary_stats = {
        'counts_mandatory': {
            'correct_fields': 0, 
            'incorrect_fields': 0, 
            'not_present_fields': 0, 
            'total_defined_mandatory_fields': 0
        },
        'counts_optional': {
            'filled_occurrences': 0, 
            'empty_occurrences': 0, 
            'total_defined_optional_fields': 0, 
            'total_possible_optional_occurrences': 0
        },
        'total_rows_in_df': total_rows,
        'collapsed_fields_count': len(collapsed_fields),
        'template_context': template_context
    }

    # Initialiseer tellers en resultaatlijsten
    for field_name in fields_config:
        # Skip ingeklapte velden in Template Generator templates
        if field_name in collapsed_fields:
            logging.debug(f"Overslaan ingeklapt veld: {field_name}")
            continue
            
        # Controleer of 'importance' bestaat en 'Verplicht' is, anders default naar optioneel
        rules = fields_config[field_name]
        is_mandatory = rules.get('importance') == 'Verplicht'
        
        if is_mandatory:
            summary_stats['counts_mandatory']['total_defined_mandatory_fields'] += 1
        else:
            summary_stats['counts_optional']['total_defined_optional_fields'] += 1

        if field_name in df.columns:
            filled_counts[field_name] = 0
            field_validation_results[field_name] = []

    # --- Hoofd loop door rijen ---
    logging.info(f"Start validatie van {total_rows} rijen...")
    for idx, row in df.iterrows():
        excel_row_num = idx + rij_offset
        row_data = row.to_dict() # Voor dependency checks

        # Valideer elk veld in de rij
        for field, rules in fields_config.items():
            # Skip ingeklapte velden in Template Generator templates
            if field in collapsed_fields:
                continue
                
            if field in df.columns:
                value = row[field]
                value_str = '' if pd.isnull(value) else str(value).strip()

                # Tel gevulde velden
                is_gevuld = value_str != '' and value_str.lower() not in invalid_values
                if is_gevuld:
                    filled_counts[field] = filled_counts.get(field, 0) + 1

                # Voer validatie uit - gebruik v20 native als rules array aanwezig is
                if "rules" in rules:
                    # v20 native validation
                    # Reference lists uit global scope halen 
                    reference_lists_data = globals().get('loaded_reference_lists', None)
                    errors = validate_field_v20_native(field, value_str if is_gevuld else value, rules, invalid_values, row_data, reference_lists_data)
                else:
                    # v18 legacy validation
                    errors = validate_field(field, value_str if is_gevuld else value, rules, invalid_values, row_data)

                # Verwerk gevonden fouten
                if errors:
                     for err in errors:
                         if err.get("message") and str(err.get("message")).strip():
                             result_item = {
                                 "Rij": excel_row_num,
                                 "GHX Kolom": field,
                                 "Supplier Kolom": clean_supplier_header(original_column_mapping.get(field, field)),
                                 "Veldwaarde": value_str, # Altijd de string waarde opslaan
                                 "Foutmelding": err.get('message', ''),
                                 "code": err.get('code', '')
                             }
                             results.append(result_item)
                             if field in field_validation_results:
                                 field_validation_results[field].append(result_item)

        # --- Red Flag Checks per rij (uit JSON config) ---
        # Support both v18 red_flags and v20 global_validations
        if "field_validations" in validation_config:
            # v20 structure - gebruik global_validations
            red_flags_config = validation_config.get("global_validations", [])
            logging.info(f"Gebruik v20 global_validations: {len(red_flags_config)} validaties gevonden")
        else:
            # v18 structure - gebruik red_flags
            red_flags_config = validation_config.get("red_flags", [])
            logging.info(f"Gebruik v18 red_flags: {len(red_flags_config)} validaties gevonden")
        for flag in red_flags_config:
            try:
                condition = flag.get("condition")
                # v20 gebruikt 'message', v18 gebruikt 'error_message'
                message = flag.get("message") or flag.get("error_message")
                flag_fields = flag.get("fields", [])

                if condition == "both_empty" or condition == "all_fields_empty":
                    # Check alleen als de velden daadwerkelijk bestaan in de dataframe
                    relevant_flag_fields = [f for f in flag_fields if f in df.columns]
                    if len(relevant_flag_fields) == len(flag_fields): # Alleen checken als alle velden aanwezig zijn
                        # Gebruik row_data (dict) voor checken, inclusief check op invalid_values
                        all_empty = all(
                            pd.isnull(row_data.get(f)) or
                            str(row_data.get(f)).strip() == '' or
                            str(row_data.get(f)).strip().lower() in invalid_values
                            for f in relevant_flag_fields
                        )
                        if all_empty:
                            # Voeg alleen toe als het bericht nog niet globaal is toegevoegd
                            existing_messages = [item["message"] for item in red_flag_messages_list]
                            if message and message not in existing_messages:
                                code = flag.get("code", "800")  # v20 en v18 ondersteuning
                                red_flag_messages_list.append({"message": message, "code": code})
                                logging.debug(f"Red flag '{condition}' getriggerd voor rij {excel_row_num}")
                
                elif condition == "uom_match" or condition == "uom_match_if_base_and_orderable":
                    # Check of UOM codes gelijk moeten zijn als base=orderable beide 1 zijn
                    required_fields = ["Is BestelbareEenheid", "Is BasisEenheid", "UOM Code Verpakkingseenheid", "UOM Code Basiseenheid"]
                    if all(f in df.columns for f in required_fields):
                        is_base = str(row_data.get("Is BasisEenheid", "")).strip()
                        is_orderable = str(row_data.get("Is BestelbareEenheid", "")).strip()
                        uom_trade = str(row_data.get("UOM Code Verpakkingseenheid", "")).strip()
                        uom_base = str(row_data.get("UOM Code Basiseenheid", "")).strip()
                        
                        if is_base == "1" and is_orderable == "1":
                            if uom_trade and uom_base and uom_trade != uom_base:
                                existing_messages = [item["message"] for item in red_flag_messages_list]
                                if message and message not in existing_messages:
                                    code = flag.get("code", "801")  # v20 en v18 ondersteuning
                                    red_flag_messages_list.append({"message": message, "code": code})
                                    logging.debug(f"Red flag UOM mismatch getriggerd voor rij {excel_row_num}: {uom_trade} != {uom_base}")
                
                elif condition == "content_match" or condition == "content_match_if_base_and_orderable":
                    # Check of Inhoud velden gelijk moeten zijn als base=orderable beide 1 zijn
                    required_fields = ["Is BestelbareEenheid", "Is BasisEenheid", "Inhoud Verpakkingseenheid", "Inhoud Basiseenheid"]
                    if all(f in df.columns for f in required_fields):
                        is_base = str(row_data.get("Is BasisEenheid", "")).strip()
                        is_orderable = str(row_data.get("Is BestelbareEenheid", "")).strip()
                        content_trade = str(row_data.get("Inhoud Verpakkingseenheid", "")).strip()
                        content_base = str(row_data.get("Inhoud Basiseenheid", "")).strip()
                        
                        if is_base == "1" and is_orderable == "1":
                            if content_trade and content_base and content_trade != content_base:
                                existing_messages = [item["message"] for item in red_flag_messages_list]
                                if message and message not in existing_messages:
                                    code = flag.get("code", "805")  # v20 en v18 ondersteuning
                                    red_flag_messages_list.append({"message": message, "code": code})
                                    logging.debug(f"Red flag Content mismatch getriggerd voor rij {excel_row_num}: {content_trade} != {content_base}")
                
                elif condition == "uom_relation" or condition == "uom_relation_conflict":
                    # Check voor UOM relatie conflicten - wordt al gehandeld door validate_uom_relationships
                    # Deze flag wordt toegevoegd in validate_uom_relationships functie als er UOM errors zijn
                    pass
                    
                elif condition == "uom_description_format" or condition == "uom_description_format_mismatch":
                    # Check wordt al gehandeld in validate_uom_relationships functie
                    # Slaat deze check over om dubbele meldingen te voorkomen
                    pass
                                    
                elif condition == "incomplete_dimensions" or condition == "incomplete_set":
                    # Check of afmetingen set compleet is
                    required_fields = ["Hoogte", "Breedte", "Diepte"]
                    available_fields = [f for f in required_fields if f in df.columns]
                    if available_fields:  # Als minimaal één afmeting veld beschikbaar is
                        values = []
                        for field in available_fields:
                            value = str(row_data.get(field, "")).strip()
                            values.append(value)
                        
                        # Tel hoeveel velden ingevuld zijn (niet leeg en niet invalid)
                        filled_count = sum(1 for v in values if v and v.lower() not in invalid_values)
                        
                        # Als er minimaal 1 ingevuld is maar niet alle beschikbare velden
                        if filled_count > 0 and filled_count < len(available_fields):
                            existing_messages = [item["message"] for item in red_flag_messages_list]
                            if message and message not in existing_messages:
                                code = flag.get("code", "804")  # v20 en v18 ondersteuning
                                red_flag_messages_list.append({"message": message, "code": code})
                                logging.debug(f"Red flag Incomplete dimensions voor rij {excel_row_num}")
                
                # Voeg hier checks toe voor andere per-rij condities indien nodig

            except Exception as e:
                logging.error(f"Red flag check error in rij {excel_row_num}: {e}")
                continue

    logging.info("Validatie per rij voltooid.")
    # --- Na de hoofd loop ---

    # 1. Roep UOM validatie aan
    logging.info("Starten UOM relatie validatie...")
    # Geef original_column_mapping mee aan UOM validatie voor betere supplier kolom info
    # Let op: validate_uom_relationships past 'results' direct aan en voegt eventueel RED FLAGS toe
    results = validate_uom_relationships(df, results, validation_config, original_column_mapping) # Pass mapping
    logging.info("UOM relatie validatie voltooid.")


    # 2. Bereken summary_stats NA ALLE validaties ---
    logging.info("Berekenen vullingspercentages...") # Hergebruik logging message, past nu beter
    for field_name, rules in fields_config.items():
        # Skip ingeklapte velden in Template Generator templates
        if field_name in collapsed_fields:
            continue
            
        is_mandatory = rules.get('importance') == 'Verplicht'

        if field_name in df.columns:
            if is_mandatory:
                # Heeft dit veld fouten?
                if not field_validation_results.get(field_name): # Geen lijst met errors voor dit veld
                    summary_stats['counts_mandatory']['correct_fields'] += 1
                else:
                    summary_stats['counts_mandatory']['incorrect_fields'] += 1
            else: # Optioneel veld
                filled_in_col = filled_counts.get(field_name, 0)
                summary_stats['counts_optional']['filled_occurrences'] += filled_in_col
                summary_stats['counts_optional']['empty_occurrences'] += (total_rows - filled_in_col)
                summary_stats['counts_optional']['total_possible_optional_occurrences'] += total_rows
        else: # Veld niet in DataFrame (niet in input Excel)
            if is_mandatory:
                summary_stats['counts_mandatory']['not_present_fields'] += 1
            # Voor optionele velden die niet aanwezig zijn, doen we niets met filled/empty occurrences

    filled_percentages = summary_stats # Wijs de berekende statistieken toe aan de return variabele

    logging.info("Verwerken Red Flag berichten...")
    # Verwerk Red Flags die door UOM validatie zijn toegevoegd aan validation_results
    validation_red_flags_dicts = [r for r in results if r.get("GHX Kolom") == "RED FLAG"]
    for flag_dict in validation_red_flags_dicts:
         msg = flag_dict.get("Foutmelding")
         existing_messages = [item["message"] for item in red_flag_messages_list]
         if msg and msg not in existing_messages:
              code = flag_dict.get("code", "724")  # UOM relaties meestal
              red_flag_messages_list.append({"message": msg, "code": code})

    # Voeg globale staffel check toe als Red Flag
    # Support both v18 and v20 structures
    flags_for_global = validation_config.get("red_flags", []) if "red_flags" in validation_config else validation_config.get("global_validations", [])
    staffel_check_flag = next((flag for flag in flags_for_global if flag.get("condition") == "has_staffel"), None)
    if staffel_check_flag:
         staffel_fields = staffel_check_flag.get("fields", [])
         # Check of een van de staffel kolommen data bevat
         has_staffel_data = any(df[f].notna().any() for f in staffel_fields if f in df.columns)
         if has_staffel_data:
              # v20 gebruikt 'message', v18 gebruikt 'error_message'
              msg = staffel_check_flag.get("message") or staffel_check_flag.get("error_message")
              existing_messages = [item["message"] for item in red_flag_messages_list]
              if msg and msg not in existing_messages:
                   code = staffel_check_flag.get("code", "850")  # Staffel code
                   red_flag_messages_list.append({"message": msg, "code": code})

    # Template check conditie: controleer of alle vereiste kolommen voor de nieuwe template aanwezig zijn
    template_check_flag = next((flag for flag in flags_for_global 
                              if flag.get("condition") in ["template_check", "template_column_missing"]), None)
    if template_check_flag:
        template_fields = template_check_flag.get("fields", [])
        # Controleer of alle template-specifieke velden aanwezig zijn in de dataframe
        template_fields_missing = [f for f in template_fields if f not in df.columns]
        # Als er velden missen, is het niet de nieuwste template
        if template_fields_missing:
            # v20 gebruikt 'message', v18 gebruikt 'error_message'
            msg = template_check_flag.get("message") or template_check_flag.get("error_message")
            existing_messages = [item["message"] for item in red_flag_messages_list]
            if msg and msg not in existing_messages:
                code = template_check_flag.get("code", "802")  # Template check code
                red_flag_messages_list.append({"message": msg, "code": code})
                logging.info(f"Red flag 'template_check' getriggerd: ontbrekende velden: {template_fields_missing}")

    # Verwijder duplicaten uit de verzamelde lijst (behoud dict structuur)
    seen_messages = set()
    unique_red_flags = []
    for item in red_flag_messages_list:
        msg = item["message"]
        if msg not in seen_messages:
            seen_messages.add(msg)
            unique_red_flags.append(item)
    red_flag_messages = unique_red_flags # Dit is de uiteindelijke lijst met {"message": str, "code": str}

    logging.info(f"Validatie analyse voltooid. Gevonden meldingen: {len(results)}, Unieke Red Flags: {len(red_flag_messages)}")

    # 3. Bereken errors_per_field (aantal unieke rijen met fouten per veld, max = gevuld)
    logging.info("Berekenen fouten per veld...")
    errors_per_field = {}
    for field_name, field_errors_list in field_validation_results.items():
        # Tel alleen unieke rijen met fouten voor dit veld
        unique_error_rows = set(res["Rij"] for res in field_errors_list)
        error_count = len(unique_error_rows)
        # Aantal fouten kan niet groter zijn dan aantal gevulde velden
        errors_per_field[field_name] = min(error_count, filled_counts.get(field_name, 0))

    # Retourneer alle berekende informatie
    return results, filled_percentages, red_flag_messages, errors_per_field


# -----------------------------
# HOOFDFUNCTIE VOOR AANROEP VANUIT STREAMLIT
# -----------------------------

def validate_pricelist(input_excel_path: str, mapping_json_path: str, validation_json_path: str, original_input_filename: str, reference_json_path: str = None) -> str | None:
    """
    Valideert een Excel prijslijst en genereert een Excel validatierapport.
    Retourneert het pad naar het rapport, of None bij een fout.
    """
    try:
        # 1. Laad configuraties
        logging.info("Laden configuratiebestanden...")
        try:
            with open(mapping_json_path, 'r', encoding='utf-8') as f:
                header_mapping_config = json.load(f)
            with open(validation_json_path, 'r', encoding='utf-8') as f:
                validation_config_raw = json.load(f)
            
            # Laad reference lists indien beschikbaar
            reference_lists = {}
            if reference_json_path and os.path.exists(reference_json_path):
                with open(reference_json_path, 'r', encoding='utf-8') as f:
                    reference_lists = json.load(f)
                logging.info("Reference lists geladen.")
            
            # Maak reference_lists beschikbaar in global scope voor v20 native validation
            global loaded_reference_lists
            loaded_reference_lists = reference_lists
            
        except FileNotFoundError as e:
            logging.error(f"Configuratiebestand niet gevonden: {e}")
            raise # Gooi error door naar Streamlit app
        except json.JSONDecodeError as e:
            logging.error(f"Fout bij lezen JSON configuratie: {e}")
            raise # Gooi error door

        # Detecteer JSON versie en normaliseer indien nodig
        json_version = detect_json_version(validation_config_raw)
        logging.info(f"JSON versie gedetecteerd: {json_version}")
        
        if json_version == "v20":
            # Converteer v20 naar v18-compatibele structuur
            validation_config = normalize_v20_to_v18_structure(validation_config_raw, reference_lists)
            logging.info("JSON v2.0 geconverteerd naar v18-compatibele structuur.")
        else:
            # Gebruik v18 structuur as-is
            validation_config = validation_config_raw
        
        # Haal mapping dictionary op
        header_mapping_dict = {k: v["alternatives"] for k, v in header_mapping_config.get("standard_headers", {}).items()}
        
        # Template-aware mandatory fields bepaling
        ghx_mandatory_fields = determine_mandatory_fields_for_template(input_excel_path)
        logging.info(f"GHX Verplichte velden geladen: {len(ghx_mandatory_fields)} velden.")
        
        # Template context extraheren voor rapportage
        template_context = None
        if has_template_generator_stamp(input_excel_path):
            template_context = extract_template_generator_context(input_excel_path)
            if template_context:
                # Log product type informatie voor toekomstige cross-validation
                product_types = template_context.get("product_type", "onbekend")
                institutions = template_context.get("institutions", [])
                logging.info(f"Product types: {product_types} (voor toekomstige UNSPSC cross-validatie)")
                if institutions:
                    logging.info(f"Instellingen: {', '.join(institutions)}")
            logging.info("Template Generator context geëxtraheerd voor rapportage")

        # 2. Lees Excel in
        logging.info(f"Lezen Excel bestand: {input_excel_path}")
        try:
            string_columns = [field for field, rules in validation_config.get("fields", {}).items()
                              if rules.get("read_as_string")]
            dtype_spec = {col: str for col in string_columns}
            
            # Voeg expliciete dtype mappings toe voor problematische kolomnamen in de originele Excel
            # Voor UNSPSC Code (zowel GHX als supplier-kolomnamen)
            dtype_spec["UNSPSC CODE (UNITED NATIONS STANDARD PRODUCTS AND SERVICES CODE)"] = str
            dtype_spec["UNSPSC Code"] = str
            
            # Voor GTIN en andere kolommen die numeriek lijken maar string moeten zijn
            dtype_spec["BARCODENUMMER (EAN/ GTIN/ HIBC)"] = str
            dtype_spec["GTIN Verpakkingseenheid"] = str
            
            # Lees Excel in met de uitgebreide dtype_spec
            df = pd.read_excel(input_excel_path, dtype=dtype_spec)
            df_original = df.copy()
            
            # DEBUG: Log de ruwe, onbewerkte headers (debug level)
            logging.debug("=== STAP 1: RUWE HEADERS ===")
            for i, col in enumerate(df.columns):
                # Toon exacte string inclusief verborgen karakters
                col_repr = repr(col)
                logging.debug(f"Ruwe header {i}: {col_repr}")
                # Toon specifiek problematische headers
                if "hoogte" in str(col).lower() or "inhoud basiseenheid" in str(col).lower():
                    logging.debug(f"PROBLEMATISCHE HEADER {i}: {col_repr}")
            
            # STAP 1.5: PREPROCESSING - Normaliseer nieuwe GHX template headers
            logging.info("Normaliseren van nieuwe GHX template headers...")
            original_columns = df.columns.tolist()
            normalized_columns = []
            
            for col in original_columns:
                # Check of dit een nieuwe template header is (bevat newlines en underscores scheidingslijn)
                if isinstance(col, str) and '\n' in col and '_' in col:
                    normalized = normalize_template_header(col)
                    if normalized != col:
                        logging.info(f"Header genormaliseerd: {repr(col[:50])}... → {repr(normalized)}")
                    normalized_columns.append(normalized)
                else:
                    normalized_columns.append(col)
            
            # Vervang de kolomnamen in de DataFrame
            df.columns = normalized_columns
            
            logging.info(f"Excel succesvol gelezen: {df.shape[0]} rijen, {df.shape[1]} kolommen.")
        except Exception as e:
            logging.error(f"Fout bij lezen Excel bestand: {e}")
            raise # Gooi error door


        # 3. Headers mappen
        logging.info("Mappen van headers...")
        df, unrecognized, duplicates, original_column_mapping = map_headers(df, header_mapping_config, return_mapping=True)
        logging.info(f"Header mapping voltooid. Onherkend: {len(unrecognized)}, Duplicaten: {len(duplicates)}.")
        if unrecognized:
             logging.warning(f"Onherkende headers gevonden: {unrecognized}")
             # Hier zou je eventueel de lijst 'unrecognized' kunnen teruggeven of loggen


        # 4. Data opschonen
        logging.info("Opschonen DataFrame...")
        df = clean_dataframe(df)
        if df.empty:
             logging.warning("DataFrame is leeg na opschonen.")
             # Overweeg hier te stoppen of een leeg rapport te maken
        else:
             logging.info(f"DataFrame opgeschoond. Resterende rijen: {len(df)}")


        # 5. Valideer data
        logging.info("Starten validatie DataFrame...")
        
        # Gebruik native v20 als beschikbaar, anders v18
        config_for_validation = validation_config_raw if json_version == "v20" else validation_config
        
        results, filled_percentages, red_flag_messages, errors_per_field = validate_dataframe(
            df, config_for_validation, original_column_mapping, template_context
        )


        # 6. Bepaal bestandsnaam en output directory voor het rapport
        report_base_name = original_input_filename # Gebruik de doorgegeven originele naam
        # Output directory strategie: submap in dezelfde map als input
        output_dir_base = os.path.dirname(input_excel_path)
        # Maak een submap 'validation_reports' als die nog niet bestaat
        report_subdir = os.path.join(output_dir_base, "validation_reports")
        os.makedirs(report_subdir, exist_ok=True)
        # Voeg timestamp toe aan de mapnaam binnen de submap
        output_dir_timestamped = os.path.join(report_subdir, "report_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        os.makedirs(output_dir_timestamped, exist_ok=True)
        logging.info(f"Output directory voor rapport: {output_dir_timestamped}")


        # 7. Genereer rapport met alle benodigde argumenten
        logging.info("Genereren validatierapport...")
        # Roep de geïmporteerde functie aan
        output_path = genereer_rapport(
            validation_results=results,
            output_dir=output_dir_timestamped,
            bestandsnaam=report_base_name, # <-- Gebruik nu report_base_name (originele naam)
            ghx_mandatory_fields=ghx_mandatory_fields,
            original_column_mapping=original_column_mapping,
            df=df,
            header_mapping=header_mapping_dict,
            df_original=df_original,
            red_flag_messages=red_flag_messages,
            errors_per_field=errors_per_field,
            JSON_CONFIG_PATH=validation_json_path, # Nodig voor laden config in rapport
            summary_data=filled_percentages,  # Doorgeven van summary_data
            validation_config=validation_config,  # Genormaliseerde config in plaats van validation_config_raw
            template_context=template_context,  # Template Generator context voor rapportage
            excel_path=input_excel_path,  # Pad naar origineel Excel bestand voor template detectie
        )

        if output_path:
            logging.info(f"Rapport succesvol gegenereerd: {output_path}")
            return output_path
        else:
            logging.error("Genereren van rapport is mislukt (genereer_rapport gaf None terug).")
            return None

    except FileNotFoundError as e:
        logging.error(f"Bestand niet gevonden tijdens validatie: {e}")
        raise
    except ImportError as e:
         logging.error(f"Import fout tijdens validatie: {e}")
         raise
    except Exception as e:
        logging.error(f"Een onverwachte fout is opgetreden tijdens validate_pricelist: {e}", exc_info=True)
        raise # Gooi de error opnieuw op zodat Streamlit het kan tonen