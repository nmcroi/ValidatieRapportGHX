"""
Excel manipulatie module voor GHX Template Generator.

Bevat functies voor het aanpassen van Excel bestanden met openpyxl.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from .engine import FieldDecision


class ExcelProcessor:
    """
    Excel bestand processor voor template aanpassingen.
    """
    
    # Standaard kleuren
    DEFAULT_MANDATORY_COLOR = "#f89a8c"  # Oranje/zalm
    DEFAULT_HIDDEN_COLOR = "#EEEEEE"     # Lichtgrijs
    
    def __init__(self, 
                 mandatory_color: str = DEFAULT_MANDATORY_COLOR,
                 hidden_color: str = DEFAULT_HIDDEN_COLOR):
        """
        Initialiseer Excel processor.
        
        Args:
            mandatory_color: Hex kleur voor verplichte velden
            hidden_color: Hex kleur voor verborgen velden (reserved voor toekomst)
        """
        self.mandatory_fill = PatternFill(
            start_color=mandatory_color.lstrip('#'),
            end_color=mandatory_color.lstrip('#'),
            fill_type="solid"
        )
        self.hidden_fill = PatternFill(
            start_color=hidden_color.lstrip('#'),
            end_color=hidden_color.lstrip('#'),
            fill_type="solid"
        )
        
        # Stijlen voor comments
        self.comment_font = Font(name="Calibri", size=9)
    
    def process_template(self, 
                        input_path: Path, 
                        output_path: Path,
                        decisions: List[FieldDecision],
                        context_dict: Dict[str, Any],
                        sheet_name: str = "Sheet1") -> None:
        """
        Verwerk template bestand en genereer aangepaste versie.
        
        Args:
            input_path: Pad naar input template
            output_path: Pad voor output bestand
            decisions: Lijst van veld beslissingen
            context_dict: Context dictionary voor stempel
            sheet_name: Naam van sheet om aan te passen
        """
        # Laad workbook
        try:
            wb = load_workbook(input_path)
        except Exception as e:
            raise ValueError(f"Kan template niet laden van {input_path}: {e}")
        
        # Controleer of sheet bestaat
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' niet gevonden in template. Beschikbare sheets: {wb.sheetnames}")
        
        ws = wb[sheet_name]
        
        # Maak beslissingen dictionary per kolom
        decisions_by_col = {d.column: d for d in decisions}
        
        # Bewaar originele kolombreedte voordat we kolommen gaan verbergen
        original_widths = self._preserve_column_widths(ws)
        
        # Pas kolommen aan
        self._apply_column_decisions(ws, decisions_by_col)
        
        # Herstel kolombreedte voor zichtbare kolommen
        self._restore_visible_column_widths(ws, decisions_by_col, original_widths)
        
        # FINALE CLEANUP: Force verborgen kolommen naar width=0 (overschrijft alles)
        for col, decision in decisions_by_col.items():
            if not decision.visible:
                ws.column_dimensions[col].width = 0
        
        # Voeg stempel toe
        self._add_stamp(wb, context_dict)
        
        # Sla op
        try:
            wb.save(output_path)
        except Exception as e:
            raise ValueError(f"Kan bestand niet opslaan naar {output_path}: {e}")
    
    def _apply_column_decisions(self, ws: Worksheet, decisions_by_col: Dict[str, FieldDecision]) -> None:
        """
        Pas kolom beslissingen toe op worksheet.
        
        Args:
            ws: Worksheet object
            decisions_by_col: Dictionary met kolom -> FieldDecision mapping
        """
        for column, decision in decisions_by_col.items():
            # Zet kolom zichtbaarheid
            ws.column_dimensions[column].hidden = not decision.visible
            
            # Skip verdere styling als kolom niet zichtbaar
            if not decision.visible:
                continue
            
            # Pas mandatory styling toe
            if decision.mandatory:
                self._apply_mandatory_styling(ws, column)
            
            # Comments uitgeschakeld - geen rode hoekjes
            # if decision.notes:
            #     self._add_header_comment(ws, column, decision.notes)
        
        # Forceer verbergen van kolommen die in beslissingen als niet-zichtbaar zijn gemarkeerd
        # Dit respecteert de context (bijv. AA/AB zijn zichtbaar in staffel context)
        permanently_hidden = [col for col, decision in decisions_by_col.items() 
                             if not decision.visible and col in ['AA', 'AB']]
        if permanently_hidden:
            self.hide_columns_permanently(ws, permanently_hidden)
            # Na enhanced hiding, zet width expliciet op 0 voor verborgen kolommen
            for col in permanently_hidden:
                ws.column_dimensions[col].width = 0
    
    def _apply_mandatory_styling(self, ws: Worksheet, column: str) -> None:
        """
        Pas mandatory styling toe op kolom header.
        Behoudt alle bestaande font eigenschappen.
        
        Args:
            ws: Worksheet object
            column: Kolom letter
        """
        header_cell = ws[f"{column}1"]
        
        # Bewaar bestaande font eigenschappen VOOR we iets wijzigen
        current_font = header_cell.font
        original_name = current_font.name if current_font else None
        original_size = current_font.size if current_font else None
        original_bold = current_font.bold if current_font else None
        original_italic = current_font.italic if current_font else None
        original_color = current_font.color if current_font else None
        original_underline = current_font.underline if current_font else None
        original_strike = current_font.strike if current_font else None
        
        # Alleen achtergrondkleur wijzigen
        header_cell.fill = self.mandatory_fill
        
        # Als er origineel font was, herstel alle eigenschappen
        if current_font:
            header_cell.font = Font(
                name=original_name,
                size=original_size,
                bold=original_bold,
                italic=original_italic,
                color=original_color,
                underline=original_underline,
                strike=original_strike
            )
    
    def _add_header_comment(self, ws: Worksheet, column: str, note_text: str) -> None:
        """
        Voeg comment toe aan kolom header.
        
        Args:
            ws: Worksheet object
            column: Kolom letter
            note_text: Comment tekst
        """
        header_cell = ws[f"{column}1"]
        
        # Maak comment object
        comment = Comment(note_text, "GHX Template Generator")
        comment.width = 300
        comment.height = 100
        
        # Voeg toe aan cel
        header_cell.comment = comment
    
    def _add_stamp(self, wb: Workbook, context_dict: Dict[str, Any]) -> None:
        """
        Voeg GHX stempel toe aan workbook.
        
        Args:
            wb: Workbook object
            context_dict: Context dictionary voor JSON embed
        """
        # Maak hidden sheet voor metadata
        self._create_metadata_sheet(wb, context_dict)
        
        # Maak named range voor compacte code
        self._create_stamp_named_range(wb, context_dict)
        
        # Voeg zichtbare configuratie code toe aan hoofdsheet
        self._add_visible_config_code(wb, context_dict)
    
    def _add_visible_config_code(self, wb: Workbook, context_dict: Dict[str, Any]) -> None:
        """
        Voeg zichtbare configuratie code toe aan cellen A4 en A5.
        
        Args:
            wb: Workbook object
            context_dict: Context dictionary met configuratie
        """
        from openpyxl.styles import PatternFill, Font
        
        # Get het eerste werkblad (Template NL)
        if "Template NL" in wb.sheetnames:
            ws = wb["Template NL"]
        else:
            ws = wb.active
        
        # Haal de configuratie code op uit de context
        config_code = context_dict.get('_config_code', 'B-UNKNOWN')
        
        # Styling: rode achtergrond, zwarte tekst
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        black_font = Font(color="000000", bold=True)
        
        # A4: Instructie tekst
        ws["A4"] = "Deze code niet verwijderen:"
        ws["A4"].fill = red_fill
        ws["A4"].font = black_font
        
        # A5: Configuratie code
        ws["A5"] = config_code
        ws["A5"].fill = red_fill
        ws["A5"].font = black_font
    
    def _create_metadata_sheet(self, wb: Workbook, context_dict: Dict[str, Any]) -> None:
        """
        Maak verborgen metadata sheet.
        
        Args:
            wb: Workbook object
            context_dict: Context dictionary
        """
        # Verwijder bestaande metadata sheet als aanwezig
        if "_GHX_META" in wb.sheetnames:
            wb.remove(wb["_GHX_META"])
        
        # Maak nieuwe sheet
        meta_ws = wb.create_sheet("_GHX_META")
        meta_ws.sheet_state = "hidden"
        
        # Schrijf prettified JSON naar A1
        import json
        json_str = json.dumps(context_dict, indent=2, ensure_ascii=False)
        meta_ws["A1"] = json_str
        
        # Voeg timestamp toe
        from datetime import datetime
        meta_ws["A2"] = f"Generated: {datetime.now().isoformat()}"
        meta_ws["A3"] = "GHX Template Generator v1.0.0"
    
    def _create_stamp_named_range(self, wb: Workbook, context_dict: Dict[str, Any]) -> None:
        """
        Maak named range voor compacte stempel code.
        
        Args:
            wb: Workbook object
            context_dict: Context dictionary
        """
        # Genereer compacte code
        from .context import Context
        
        # Reconstruct context object om preset code te genereren
        try:
            context_obj = Context(**context_dict)
            preset_code = context_obj.get_preset_code()
        except Exception:
            # Fallback als context reconstructie faalt
            preset_code = "CUSTOM-UNKNOWN"
        
        # Verwijder bestaande named range
        if "GHX_STAMP" in wb.defined_names:
            del wb.defined_names["GHX_STAMP"]
        
        # Voeg compacte code toe aan metadata sheet
        if "_GHX_META" in wb.sheetnames:
            meta_ws = wb["_GHX_META"]
            meta_ws["B1"] = preset_code
            
            # Maak named range die verwijst naar deze cel
            from openpyxl.workbook.defined_name import DefinedName
            defn = DefinedName("GHX_STAMP", attr_text=f"_GHX_META!$B$1")
            wb.defined_names["GHX_STAMP"] = defn
    
    def extract_stamp(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """
        Extraheer GHX stempel van bestaand bestand.
        
        Args:
            file_path: Pad naar Excel bestand
            
        Returns:
            Context dictionary of None als geen stempel gevonden
        """
        try:
            wb = load_workbook(file_path)
            
            # Probeer metadata sheet te lezen
            if "_GHX_META" in wb.sheetnames:
                meta_ws = wb["_GHX_META"]
                json_str = meta_ws["A1"].value
                
                if json_str:
                    import json
                    return json.loads(json_str)
            
            return None
            
        except Exception as e:
            print(f"Waarschuwing: Kon stempel niet extraheren van {file_path}: {e}")
            return None
    
    def validate_template(self, file_path: Path, required_sheets: List[str] = None) -> List[str]:
        """
        Valideer template bestand structuur.
        
        Args:
            file_path: Pad naar template bestand
            required_sheets: Lijst van vereiste sheet namen
            
        Returns:
            Lijst van validatie fouten (leeg als geldig)
        """
        errors = []
        required_sheets = required_sheets or ["Sheet1"]
        
        try:
            wb = load_workbook(file_path, read_only=True)
            
            # Controleer vereiste sheets
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    errors.append(f"Vereiste sheet '{sheet_name}' niet gevonden")
            
            # Controleer of Sheet1 headers heeft
            if "Sheet1" in wb.sheetnames:
                ws = wb["Sheet1"]
                if ws.max_row < 1:
                    errors.append("Sheet1 bevat geen data")
                elif not any(ws[f"{chr(65+i)}1"].value for i in range(26)):  # A1-Z1
                    errors.append("Sheet1 lijkt geen headers te hebben in rij 1")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Kan template bestand niet lezen: {e}")
        
        return errors
    
    def hide_columns_permanently(self, 
                                worksheet: Worksheet,
                                columns_to_hide: List[str],
                                method: str = "all_methods") -> None:
        """
        Verberg kolommen permanent met verschillende methodes voor maximale compatibiliteit.
        
        Args:
            worksheet: Excel worksheet
            columns_to_hide: Lijst van kolom letters (bijv. ['AA', 'AB'])
            method: Verstop methode ('all_methods' aanbevolen voor beste compatibiliteit)
        """
        try:
            from .column_hiding import ColumnHider, HideMethod
            
            hider = ColumnHider()
            method_enum = HideMethod(method)
            
            result = hider.hide_columns(worksheet, columns_to_hide, method_enum)
            
            if result['errors']:
                print(f"Waarschuwingen bij kolom verbergen: {result['errors']}")
            else:
                print(f"✅ Kolommen {columns_to_hide} succesvol verborgen met methode '{method}'")
                
        except ImportError:
            # Fallback als enhanced_column_hiding niet beschikbaar is
            print("⚠️ Enhanced column hiding niet beschikbaar, gebruik basis methode")
            for col in columns_to_hide:
                try:
                    worksheet.column_dimensions[col].hidden = True
                    worksheet.column_dimensions[col].width = 0
                    print(f"✅ Kolom {col} verborgen (basis methode)")
                except Exception as e:
                    print(f"❌ Fout bij verbergen kolom {col}: {e}")
            
            # FORCE width naar 0 voor verborgen kolommen - overschrijf enhanced hiding
            for col in columns_to_hide:
                try:
                    worksheet.column_dimensions[col].width = 0
                    print(f"✅ Kolom {col} width geforceerd naar 0")
                except Exception as e:
                    print(f"❌ Fout bij width forceren {col}: {e}")
        
        except Exception as e:
            print(f"❌ Fout bij permanent verbergen kolommen: {e}")
    
    def _preserve_column_widths(self, worksheet: Worksheet) -> Dict[str, float]:
        """
        Zet consistente kolombreedte volgens GHX standaard.
        
        Args:
            worksheet: Excel worksheet
            
        Returns:
            Dictionary met kolom -> breedte mapping
        """
        widths = {}
        
        # GHX standaard kolombreedte
        # Kolom A = 50, alle andere zichtbare kolommen = 30.00 (185 pixels)
        for i in range(1, 105):  # Tot kolom CZ
            col_letter = self._number_to_column(i)
            if col_letter == 'A':
                widths[col_letter] = 50.0  # Eerste kolom breder
            else:
                widths[col_letter] = 30.0  # Standaard breedte (185 pixels)
        
        return widths
    
    def _restore_visible_column_widths(self, 
                                      worksheet: Worksheet, 
                                      decisions_by_col: Dict[str, 'FieldDecision'],
                                      original_widths: Dict[str, float]) -> None:
        """
        Zet correcte kolombreedte: zichtbare kolommen krijgen standaard breedte, verborgen krijgen 0.
        
        Args:
            worksheet: Excel worksheet
            decisions_by_col: Kolom beslissingen
            original_widths: Standaard kolombreedte mapping
        """
        for col_letter, width in original_widths.items():
            # Check of deze kolom zichtbaar moet zijn
            if col_letter in decisions_by_col:
                decision = decisions_by_col[col_letter]
                if decision.visible:
                    # Zichtbare kolom: gebruik standaard breedte
                    worksheet.column_dimensions[col_letter].width = width
                else:
                    # Verborgen kolom: width = 0 EN zet properties die Excel zou kunnen overschrijven
                    col_dim = worksheet.column_dimensions[col_letter]
                    col_dim.width = 0
                    col_dim.hidden = True
                    # Probeer autosize uit te schakelen
                    if hasattr(col_dim, 'auto_size'):
                        col_dim.auto_size = False
                    if hasattr(col_dim, 'bestFit'):
                        col_dim.bestFit = False
            else:
                # Voor kolommen die niet in beslissingen staan (meestal zichtbaar), gebruik standaard
                worksheet.column_dimensions[col_letter].width = width
    
    @staticmethod
    def _number_to_column(n: int) -> str:
        """Convert nummer naar Excel kolom letter (1=A, 2=B, etc.)"""
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + n % 26) + result
            n //= 26
        return result
