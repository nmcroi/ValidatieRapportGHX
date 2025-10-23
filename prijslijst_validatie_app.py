# prijslijst_validatie_app.py

import streamlit as st
import pandas as pd
import os
import tempfile # Nodig om geupload bestand tijdelijk op te slaan
import logging # Om logging uit de tool te zien (optioneel)
import io
import zipfile
import time

# Importeer de hoofdfunctie uit price_tool.py (met volledige rapport functionaliteit)
try:
    from validator.price_tool import validate_pricelist
except ImportError as e:
    st.error(f"Fout bij importeren validatiemodule: {e}")
    st.error("Zorg ervoor dat de map 'validator' bestaat met daarin price_tool.py en rapport_utils.py.")
    st.stop() # Stop de app als de import faalt

# --- Basis Configuratie ---
# Pad naar de configuratiebestanden (pas aan indien nodig)
# Gaat er hier vanuit dat ze in dezelfde map staan als de app
MAPPING_JSON = "header_mapping.json"
VALIDATION_JSON = "field_validation_v20.json"
REFERENCE_JSON = "reference_lists.json"

# Logging configureren (optioneel, toont logs in console waar Streamlit draait)
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Streamlit UI Opzet ---
st.set_page_config(layout="wide") # Gebruik meer schermbreedte
st.title("✅ GHX Prijslijst Validatie Tool - BUG GEFIXED!")
st.markdown("🎯 **VERSIE MET WERKENDE UOM CONFLICT LOGIC** - Upload hier de GHX prijslijst template (.xlsx of .xls) om deze te valideren.")

# Controleer of configuratiebestanden bestaan
if not os.path.exists(MAPPING_JSON):
    st.error(f"Mapping configuratiebestand niet gevonden: {MAPPING_JSON}")
    st.stop()
if not os.path.exists(VALIDATION_JSON):
    st.error(f"Validatie configuratiebestand niet gevonden: {VALIDATION_JSON}")
    st.stop()
if not os.path.exists(REFERENCE_JSON):
    st.error(f"Reference lists bestand niet gevonden: {REFERENCE_JSON}")
    st.stop()

# Initialiseer session state voor rapport data als deze nog niet bestaat
if 'report_data' not in st.session_state:
    st.session_state.report_data = {}

# Initialiseer session state voor Quick Mode bestanden
if 'quick_mode_files' not in st.session_state:
    st.session_state.quick_mode_files = {}

# File uploader - Accept multiple files
uploaded_files = st.file_uploader("Kies een of meerdere Excel-bestanden", type=["xlsx", "xls"], accept_multiple_files=True)

# Check if files have been uploaded
if uploaded_files:
    st.info(f"{len(uploaded_files)} bestand(en) geselecteerd.")

    # Knop om validatie te starten
    if st.button(f"Start Validatie & Genereer Rapporten voor {len(uploaded_files)} bestand(en)"):
        # Reset/clear previous report data when starting new validation
        st.session_state.report_data = {}
        st.session_state.quick_mode_files = {}
        # Loop through each uploaded file
        MAX_FILE_SIZE_MB = 50
        MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

        for uploaded_file in uploaded_files:
            # Check file size before processing
            if uploaded_file.size > MAX_FILE_SIZE_BYTES:
                file_size_mb = uploaded_file.size / (1024 * 1024)
                st.error(
                    f"Bestand '{uploaded_file.name}' ({file_size_mb:.1f}MB) is te groot. "
                    f"Maximaal toegestane online validatiegrootte is {MAX_FILE_SIZE_MB}MB. "
                    f"Neem contact op met Niels Croiset om dit bestand lokaal te laten valideren."
                )
                st.session_state.report_data[uploaded_file.name] = {'error': 'file_too_large'} # Markeer als error
                continue  # Skip to the next file

            st.markdown("***") # Separator line
            st.subheader(f"Bezig met verwerken: {uploaded_file.name}")
            original_filename = uploaded_file.name
            temp_file_path = None # Reset for each file

            try:
                # Create a temporary file for THIS file
                start_time = time.time()
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                    tmp_file.write(uploaded_file.getvalue())
                    temp_file_path = tmp_file.name
                end_time = time.time()
                logging.info(f"⏱️  Tijdelijk bestand opslaan: {end_time-start_time:.2f} sec voor {len(uploaded_file.getvalue())/1024/1024:.1f} MB")

                # Template detectie gebeurt in de achtergrond tijdens validate_pricelist()
                # Geen UI output nodig - dit is alleen een testomgeving

                # Initialiseer total_rows voor het geval er een exception optreedt
                total_rows = None
                max_rows_param = None
                
                # Slimme row counting - vermijd trage openpyxl voor grote bestanden
                try:
                    import pandas as pd
                    
                    # Stap 1: Probeer eerst 5001 regels te lezen voor snelle detectie
                    df_check = pd.read_excel(temp_file_path, nrows=5001)
                    actual_rows = len(df_check)
                    
                    if actual_rows <= 5000:
                        # Bestand heeft 5000 of minder regels - we hebben het exacte aantal
                        total_rows = actual_rows
                        max_rows_param = None  # Valideer alles
                        logging.info(f"Bestand heeft {total_rows} rijen data (klein bestand).")
                        if total_rows > 0:
                            st.info(f"📊 Bestand heeft {total_rows} rijen. Volledige validatie wordt uitgevoerd.")
                    else:
                        # Bestand heeft meer dan 5000 regels
                        # Optioneel: probeer exact aantal te bepalen met openpyxl (kan traag zijn)
                        quick_count_only = True  # Zet op False als je exact aantal wilt voor grote bestanden
                        
                        # PERFORMANCE BOOST: Maak snel tijdelijk bestand voor zeer grote bestanden
                        original_temp_path = temp_file_path
                        quick_temp_path = None
                        
                        try:
                            # Test performance: is origineel bestand erg traag?
                            start_time = time.time()
                            test_df = pd.read_excel(temp_file_path, nrows=10)  # Test read 10 rows
                            test_time = time.time() - start_time
                            
                            if test_time > 0.5:  # Als lezen van 10 rijen >0.5 sec duurt, optimaliseer
                                import tempfile
                                
                                with st.spinner("⚡ Optimaliseren voor Quick Mode..."):
                                    # Maak snel tijdelijk bestand met alleen eerste 5001 rijen
                                    quick_temp_fd, quick_temp_path = tempfile.mkstemp(suffix='.xlsx')
                                    os.close(quick_temp_fd)
                                    
                                    # Gebruik de al gelezen 5001 rijen data
                                    df_check.to_excel(quick_temp_path, index=False)
                                    temp_file_path = quick_temp_path
                                    
                                    logging.info(f"✅ Quick Mode optimalisatie: {test_time:.2f}s → snel tijdelijk bestand")
                                    st.success("⚡ Bestand geoptimaliseerd voor snelle verwerking!")
                        
                        except Exception as opt_e:
                            logging.warning(f"Quick Mode optimalisatie gefaald: {opt_e}")
                        
                        if quick_count_only:
                            total_rows = "5000+"  # Voor rapport weergave
                            st.info(f"📊 Bestand heeft meer dan 5000 rijen. Quick Validatie: eerste 5000 rijen worden gevalideerd.")
                        else:
                            # Alleen als echt nodig: tel exact met openpyxl
                            with st.spinner("Exact aantal rijen wordt geteld..."):
                                import openpyxl
                                wb = openpyxl.load_workbook(original_temp_path, read_only=True)  # Gebruik origineel voor telling
                                ws = wb.active
                                total_rows = ws.max_row - 1  # -1 voor header
                                wb.close()
                            st.info(f"📊 Bestand heeft {total_rows:,} rijen. Quick Validatie: eerste 5000 rijen worden gevalideerd.")
                        
                        max_rows_param = 5000
                        logging.info(f"Bestand heeft 5000+ rijen data (groot bestand). Quick Mode actief.")
                        
                        # Sla Quick Mode info op in session state
                        st.session_state.quick_mode_files[original_filename] = {
                            'total_rows': total_rows,
                            'file_data': uploaded_file.getvalue()
                        }
                        
                except Exception as count_error:
                    st.warning(f"Kon aantal rijen niet tellen: {count_error}. Valideer volledig bestand.")
                    total_rows = None
                    max_rows_param = None

                # Start validation with spinner
                with st.spinner(f"Validatie en rapportage voor '{original_filename}' bezig..."):
                    logging.info(f"Aanroepen validate_pricelist voor {original_filename}...")
                    report_path = validate_pricelist(
                        input_excel_path=temp_file_path,
                        mapping_json_path=MAPPING_JSON,
                        validation_json_path=VALIDATION_JSON,
                        original_input_filename=original_filename,
                        reference_json_path=REFERENCE_JSON,
                        max_rows=max_rows_param,
                        total_rows=total_rows,
                    )
                    logging.info(f"validate_pricelist voltooid voor {original_filename}.")

                # Remove the temporary files after validation is done
                if os.path.exists(temp_file_path):
                     os.remove(temp_file_path)
                # Cleanup Quick Mode tijdelijk bestand als gebruikt
                if 'quick_temp_path' in locals() and quick_temp_path and os.path.exists(quick_temp_path):
                    os.remove(quick_temp_path)
                    logging.info(f"✅ Quick Mode tijdelijk bestand opgeruimd: {quick_temp_path}")

                # Show result and download button for THIS file
                if report_path and os.path.exists(report_path):
                    st.success(f"Validatierapport voor '{original_filename}' succesvol gegenereerd!")
                    try:
                        with open(report_path, "rb") as fp:
                            report_bytes = fp.read()
                        download_filename = os.path.basename(report_path)
                        # IMPORTANT: Use a unique key for each download button
                        st.download_button(
                            label=f"Download Rapport voor '{original_filename}'",
                            data=report_bytes,
                            file_name=download_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_btn_{original_filename}" # Key for button itself (optional but good practice)
                        )
                        
                        # Voeg "Volledige Validatie" knop toe voor Quick Mode bestanden
                        if max_rows_param is not None and original_filename in st.session_state.quick_mode_files:
                            stored_total = st.session_state.quick_mode_files[original_filename]['total_rows']
                            if stored_total == '5000+':
                                button_text = f"🔍 Volledige Validatie voor '{original_filename}' (5000+ rijen)"
                                info_text = "Start volledige validatie van alle rijen..."
                            else:
                                button_text = f"🔍 Volledige Validatie voor '{original_filename}' ({stored_total:,} rijen)"
                                info_text = f"Start volledige validatie van alle {stored_total:,} rijen..."
                            
                            if st.button(button_text, key=f"full_validation_{original_filename}"):
                                st.markdown("---")
                                st.subheader(f"Volledige Validatie: {original_filename}")
                                st.info(info_text)
                                
                                # Maak opnieuw een tijdelijk bestand van de opgeslagen data
                                full_temp_file_path = None
                                try:
                                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                                        tmp_file.write(st.session_state.quick_mode_files[original_filename]['file_data'])
                                        full_temp_file_path = tmp_file.name
                                    
                                    # Bij volledige validatie, tel het exacte aantal als we dat nog niet wisten
                                    actual_total_rows = stored_total
                                    if stored_total == '5000+':
                                        with st.spinner("Exact aantal rijen wordt geteld voor volledige validatie..."):
                                            import openpyxl
                                            wb = openpyxl.load_workbook(full_temp_file_path, read_only=True)
                                            ws = wb.active
                                            actual_total_rows = ws.max_row - 1  # -1 voor header
                                            wb.close()
                                            st.info(f"📊 Bestand heeft {actual_total_rows:,} rijen. Volledige validatie wordt uitgevoerd...")
                                    
                                    # Start volledige validatie (ZONDER max_rows parameter)
                                    with st.spinner(f"Volledige validatie voor '{original_filename}' bezig..."):
                                        full_report_path = validate_pricelist(
                                            input_excel_path=full_temp_file_path,
                                            mapping_json_path=MAPPING_JSON,
                                            validation_json_path=VALIDATION_JSON,
                                            original_input_filename=original_filename,
                                            reference_json_path=REFERENCE_JSON,
                                            max_rows=None,  # VOLLEDIGE validatie
                                            total_rows=actual_total_rows,
                                        )
                                    
                                    # Ruim tijdelijk bestand op
                                    if os.path.exists(full_temp_file_path):
                                        os.remove(full_temp_file_path)
                                    
                                    # Toon resultaat volledige validatie
                                    if full_report_path and os.path.exists(full_report_path):
                                        st.success(f"Volledige validatie voor '{original_filename}' succesvol voltooid!")
                                        with open(full_report_path, "rb") as fp:
                                            full_report_bytes = fp.read()
                                        full_download_filename = os.path.basename(full_report_path)
                                        
                                        st.download_button(
                                            label=f"Download Volledig Rapport voor '{original_filename}'",
                                            data=full_report_bytes,
                                            file_name=full_download_filename,
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            key=f"full_download_btn_{original_filename}"
                                        )
                                        
                                        # Update session state met volledig rapport
                                        st.session_state.report_data[f"{original_filename}_VOLLEDIG"] = {
                                            'bytes': full_report_bytes,
                                            'filename': full_download_filename
                                        }
                                    else:
                                        st.error(f"Volledige validatie voor '{original_filename}' mislukt.")
                                        
                                except Exception as full_val_err:
                                    st.error(f"Fout bij volledige validatie: {full_val_err}")
                                    if full_temp_file_path and os.path.exists(full_temp_file_path):
                                        os.remove(full_temp_file_path)
                        
                        # Store report data in session state
                        st.session_state.report_data[original_filename] = {
                            'bytes': report_bytes,
                            'filename': download_filename
                        }
                    except Exception as read_err:
                         st.error(f"Kon rapport '{download_filename}' niet lezen voor download: {read_err}")
                else:
                    st.error(f"Genereren van rapport voor '{original_filename}' mislukt.")

            except FileNotFoundError as e:
                 st.error(f"Bestandsfout bij '{original_filename}': {e}")
            except ImportError as e:
                 st.error(f"Import fout bij '{original_filename}': {e}.")
            except Exception as e:
                st.error(f"Onverwachte fout bij '{original_filename}': {e}")
            finally:
                # Ensure the temporary file for THIS iteration is cleaned up
                 if temp_file_path and os.path.exists(temp_file_path):
                     try:
                         os.remove(temp_file_path)
                         logging.info(f"Tijdelijk bestand {temp_file_path} opgeruimd.")
                     except Exception as del_err:
                         logging.error(f"Kon tijdelijk bestand {temp_file_path} niet verwijderen: {del_err}")

        st.markdown("***") # Separator after all files
        st.success("Alle geselecteerde bestanden zijn beoordeeld.") # Aangepast omdat sommige overgeslagen kunnen zijn

# --- Display Download Buttons from Session State (outside the main button click logic) ---
if st.session_state.report_data:
    st.markdown("### Beschikbare Rapporten:")
    for original_filename, data in st.session_state.report_data.items():
        if data.get('error') == 'file_too_large':
            st.warning(f"Bestand '{original_filename}' was te groot en is niet verwerkt.")
            continue # Sla downloadknop over voor te grote bestanden
        
        # Controleer of dit een _VOLLEDIG rapport is
        if not original_filename.endswith('_VOLLEDIG'):
            # Normale download knop
            st.download_button(
                label=f"Download Rapport voor '{original_filename}'",
                data=data['bytes'],
                file_name=data['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{original_filename}" # Unique key per button
            )
            
            # Voeg "Volledige Validatie" knop toe als dit een Quick Mode bestand was
            if original_filename in st.session_state.quick_mode_files:
                quick_info = st.session_state.quick_mode_files[original_filename]
                total_rows_display = quick_info['total_rows']
                if total_rows_display == '5000+':
                    button_text = f"🔍 Volledige Validatie voor '{original_filename}' (5000+ rijen)"
                    info_text = "Start volledige validatie van alle rijen..."
                else:
                    button_text = f"🔍 Volledige Validatie voor '{original_filename}' ({total_rows_display:,} rijen)"
                    info_text = f"Start volledige validatie van alle {total_rows_display:,} rijen..."
                
                if st.button(button_text, key=f"persistent_full_validation_{original_filename}"):
                    st.markdown("---")
                    st.subheader(f"Volledige Validatie: {original_filename}")
                    st.info(info_text)
                    
                    # Maak opnieuw een tijdelijk bestand van de opgeslagen data
                    full_temp_file_path = None
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                            tmp_file.write(quick_info['file_data'])
                            full_temp_file_path = tmp_file.name
                        
                        # Bij volledige validatie, tel het exacte aantal als we dat nog niet wisten
                        actual_total_rows = quick_info['total_rows']
                        if actual_total_rows == '5000+':
                            with st.spinner("Exact aantal rijen wordt geteld voor volledige validatie..."):
                                import openpyxl
                                wb = openpyxl.load_workbook(full_temp_file_path, read_only=True)
                                ws = wb.active
                                actual_total_rows = ws.max_row - 1  # -1 voor header
                                wb.close()
                                st.info(f"📊 Bestand heeft {actual_total_rows:,} rijen. Volledige validatie wordt uitgevoerd...")
                        
                        # Start volledige validatie (ZONDER max_rows parameter)
                        with st.spinner(f"Volledige validatie voor '{original_filename}' bezig..."):
                            from validator.price_tool import validate_pricelist
                            full_report_path = validate_pricelist(
                                input_excel_path=full_temp_file_path,
                                mapping_json_path=MAPPING_JSON,
                                validation_json_path=VALIDATION_JSON,
                                original_input_filename=original_filename,
                                reference_json_path=REFERENCE_JSON,
                                max_rows=None,  # VOLLEDIGE validatie
                                total_rows=actual_total_rows,
                            )
                        
                        # Ruim tijdelijk bestand op
                        if os.path.exists(full_temp_file_path):
                            os.remove(full_temp_file_path)
                        
                        # Toon resultaat volledige validatie
                        if full_report_path and os.path.exists(full_report_path):
                            st.success(f"Volledige validatie voor '{original_filename}' succesvol voltooid!")
                            with open(full_report_path, "rb") as fp:
                                full_report_bytes = fp.read()
                            full_download_filename = os.path.basename(full_report_path)
                            
                            st.download_button(
                                label=f"Download Volledig Rapport voor '{original_filename}'",
                                data=full_report_bytes,
                                file_name=full_download_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"persistent_full_download_btn_{original_filename}"
                            )
                            
                            # Update session state met volledig rapport
                            st.session_state.report_data[f"{original_filename}_VOLLEDIG"] = {
                                'bytes': full_report_bytes,
                                'filename': full_download_filename
                            }
                            
                            # Toon direct de download knop voor het volledige rapport
                            st.markdown("---")
                            st.success(f"✅ Volledig rapport klaar voor download!")
                            st.download_button(
                                label=f"📥 Download Volledig Rapport voor '{original_filename}'",
                                data=full_report_bytes,
                                file_name=full_download_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"immediate_full_download_{original_filename}"
                            )
                        else:
                            st.error(f"Volledige validatie voor '{original_filename}' mislukt.")
                            
                    except Exception as full_val_err:
                        st.error(f"Fout bij volledige validatie: {full_val_err}")
                        if full_temp_file_path and os.path.exists(full_temp_file_path):
                            os.remove(full_temp_file_path)
        else:
            # Volledig rapport download knop
            base_filename = original_filename.replace('_VOLLEDIG', '')
            st.download_button(
                label=f"Download Volledig Rapport voor '{base_filename}'",
                data=data['bytes'],
                file_name=data['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{original_filename}" # Unique key per button
            )

    # Add Download All button if more than one report exists
    if len(st.session_state.report_data) > 1:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for original_filename, data in st.session_state.report_data.items():
                if data.get('error') == 'file_too_large':
                    continue # Voeg geen te grote bestanden toe aan zip
                # Use the generated report filename inside the zip
                zip_file.writestr(data['filename'], data['bytes'])

        st.download_button(
            label="Download Alle Rapporten (.zip)",
            data=zip_buffer.getvalue(),
            file_name="validatie_rapporten.zip",
            mime="application/zip",
            key="download_all_zip"
        )

else:
    st.info("Wacht op upload van een of meerdere Excel-bestanden.")

st.markdown("---") # Een horizontale lijn voor scheiding
st.markdown("© 2025 Created by Niels Croiset | GHX | Nightstory | V270525")